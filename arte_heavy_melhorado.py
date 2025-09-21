import pandas as pd
import google.generativeai as genai
import google.api_core.exceptions as google_exceptions
import os
from dotenv import load_dotenv
import json
import time
import logging
from openpyxl.styles import PatternFill
import re
from datetime import datetime
import colorsys

# ======================================================================
# CONFIGURAÇÕES E CONSTANTES
# ======================================================================

# --- File Paths ---
BASE_DIR = r"C:\Users\pietr\OneDrive\.vscode\arte_"
CAMINHO_EDITAL = os.path.join(BASE_DIR, "DOWNLOADS", "master.xlsx")
CAMINHO_BASE = os.path.join(BASE_DIR, "DOWNLOADS", "RESULTADO_metadados", "categoria_QWEN.xlsx")
CAMINHO_SAIDA = os.path.join(BASE_DIR, "DOWNLOADS", "master_heavy.xlsx")
CAMINHO_HEAVY_EXISTENTE = CAMINHO_SAIDA

# --- Financial Parameters ---
PROFIT_MARGIN = 0.53  # MARGEM DE LUCRO 
INITIAL_PRICE_FILTER_PERCENTAGE = 0.60  # FILTRO DE PREÇO DOS PRODUTOS NA BASE

# --- AI Model Configuration ---
LLM_MODELS_FALLBACK = [
    "gemini-2.5-flash",
    "gemini-2.5-flash-lite",
    "gemini-2.0-flash-lite",
    "gemini-2.0-flash",
    "gemini-2.5-pro",
    "gemini-1.5-flash",  
]

# --- Categorization Keywords ------------------------------------
CATEGORIZATION_KEYWORDS = {
    'ACESSORIO_CORDA': ['arco','cavalete','corda','kit nut','kit rastilho'],
    'ACESSORIO_GERAL': ['bag','banco','carrilho prancha','estante de partitura','suporte'],
    'ACESSORIO_PERCURSSAO': ['baqueta','carrilhão','esteira','Máquina de Hi Hat','Pad para Bumbo','parafuso','pedal de bumbo','pele','prato','sino','talabarte','triângulo'],
    'ACESSORIO_SOPRO': ['graxa','oleo lubrificante','palheta de saxofone/clarinete'],
    'EQUIPAMENTO_AUDIO': ['fone de ouvido','globo microfone','Interface de guitarra','pedal','mesa de som','microfone'],
    'EQUIPAMENTO_CABO': ['cabo CFTV', 'cabo de rede', 'Medusa', 'switch', 'cabo_musical'],
    'EQUIPAMENTO_SOM': ['amplificador','caixa de som','cubo para guitarra'],
    "INSTRUMENTO_CORDA": ["violino","viola","violão","guitarra","baixo","violoncelo"],
    "INSTRUMENTO_PERCUSSAO": ["afuché","bateria","bombo","bumbo","caixa de guerra","caixa tenor","ganza","pandeiro","quadriton","reco reco","surdo","tambor","tarol","timbales"],
    "INSTRUMENTO_SOPRO": ["trompete","bombardino","trompa","trombone","tuba","sousafone","clarinete","saxofone","flauta","tuba bombardão","flugelhorn","euphonium"],
    "INSTRUMENTO_TECLAS": ["piano","teclado digital","glockenspiel","metalofone"],
    "INFORMATICA" :  ["projetor", "switch", "fonte energia", "drone"]
}

# --- Specs by Category ---
SPECS_BY_CATEGORY = {
    'microfone': {
        'required': ['tipo', 'padrao_polar', 'resposta_frequencia'],
        'weights': {'tipo': 0.4, 'padrao_polar': 0.3, 'resposta_frequencia': 0.3}
    },
    # Adicione mais categorias conforme necessário
}

# --- Logging Configuration ---
LOG_FILE = os.path.join(BASE_DIR, "LOGS", "arte_heavy.log")
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def load_categorized_products(file_path):
    """Carrega produtos categorizados de uma planilha Excel."""
    try:
        df = pd.read_excel(file_path)
        return df.set_index('ID_PRODUTO')[['categoria_principal', 'subcategoria']].to_dict('index')
    except FileNotFoundError:
        print(f"ERROR: Could not load categorized products file: {file_path}")
        return {}

# ============================================================ 
# FUNÇÕES DE SUPORTE
# ============================================================ 

def gerar_conteudo_com_fallback(prompt: str, modelos: list[str]) -> str | None:
    """
    Tenta gerar conteúdo usando uma lista de modelos em ordem de preferência.
    Se um modelo falhar por cota (ResourceExhausted), tenta o próximo.
    """
    for nome_modelo in modelos:
        try:
            print(f"   - Tentando chamada à API com o modelo: {nome_modelo}...")
            model = genai.GenerativeModel(nome_modelo)
            response = model.generate_content(prompt)

            if not response.parts:
                finish_reason = response.candidates[0].finish_reason.name if response.candidates else 'N/A'
                print(f"   - ❌ A GERAÇÃO RETORNOU VAZIA. Motivo: {finish_reason}. Isso pode ser causado por filtros de segurança.")
                return None
            print(f"   - Sucesso com o modelo '{nome_modelo}'.")
            return response.text
        except google_exceptions.ResourceExhausted as e:
            print(f"- Cota excedida para o modelo '{nome_modelo}'. Tentando o próximo da lista.")
            time.sleep(5)
            continue
        except Exception as e:
            print(f"   - ❌ Erro inesperado com o modelo '{nome_modelo}': {e}")
            return None
    print("   - ❌ FALHA TOTAL: Todos os modelos na lista de fallback falharam.")
    return None

def get_item_classification(description: str, reference: str, categories_with_subcategories: dict) -> dict | None:
    """Usa o modelo de IA para classificar o item, retornando categoria e subcategoria."""
    print("- Asking AI for item classification (category and subcategory)...")

    prompt = f"""<identidade>Você é um especialista de almoxarifado e perito em catalogação de produtos, com base nessas informações.</identidade>

<objetivo>
Sua tarefa é classificar o item a seguir, identificando sua `categoria_principal` e `subcategoria` com base na estrutura fornecida.
- A `categoria_principal` DEVE ser uma das chaves da estrutura abaixo.
- A `subcategoria` DEVE ser um dos valores da lista associada à categoria principal escolhida.
Responda APENAS com um objeto JSON.
</objetivo>

<item_a_ser_classificado>
Descrição: {description}
Referência: {reference}
</item_a_ser_classificado>

<estrutura_de_categorias_e_subcategorias_permitidas>
{json.dumps(categories_with_subcategories, indent=2, ensure_ascii=False)}
</estrutura_de_categorias_e_subcategorias_permitidas>

<formato_saida>
{{
  "categoria_principal": "NOME_DA_CATEGORIA_PRINCIPAL",
  "subcategoria": "NOME_DA_SUBCATEGORIA_ESPECIFICA"
}}
</formato_saida>

JSON:
"""

    response_text = gerar_conteudo_com_fallback(prompt, LLM_MODELS_FALLBACK)
    if response_text:
        try:
            cleaned_response = response_text.strip().replace("```json", "").replace("```", "")
            classification = json.loads(cleaned_response)
            
            if 'categoria_principal' in classification and 'subcategoria' in classification:
                if classification['categoria_principal'] in categories_with_subcategories:
                    return classification
                else:
                    print(f"   - WARNING: AI returned an invalid main category: '{classification['categoria_principal']}'.")
                    return None
            else:
                print("   - WARNING: AI response is missing required keys ('categoria_principal', 'subcategoria').")
                return None

        except json.JSONDecodeError as e:
            print(f"   - ERROR decoding JSON from AI for classification: {e}")
            return None
    else:
        print("   - ERROR: AI call for classification failed for all models.")
        return None

def extract_edital_specs(descricao: str, referencia: str, specs_config: dict) -> dict | None:
    specs_config_json = json.dumps(specs_config, ensure_ascii=False, indent=2)
    prompt = f"""
Você é um especialista em catalogação técnica de produtos de áudio e música. Seu objetivo é analisar o item do edital e extrair metadados estruturados.

**TAREFA:**
1. **Identifique a `categoria_primaria`**: Determine a categoria de mais alto nível do item. Deve ser um termo simples e direto.
   Exemplos: "Microfone", "Guitarra", "Mesa de Som".

2. **Extraia os `atributos` técnicos essenciais**: Com base na estrutura de specs fornecida, extraia os 'required' como críticos e outros como opcionais. Extraia de 4 a 7 atributos no total.

<estrutura_specs>
{specs_config_json}
</estrutura_specs>

**ITEM DO EDITAL:**
Descrição: {descricao}
Referência: {referencia}

**SAÍDA (APENAS JSON):**
{{
  "categoria_primaria": "<categoria>",
  "atributos": {{
    "atributo1": "<valor>",
    ...
  }}
}}
"""
    response_text = gerar_conteudo_com_fallback(prompt, LLM_MODELS_FALLBACK)
    if response_text:
        try:
            cleaned_response = response_text.strip().replace("```json", "").replace("```", "")
            return json.loads(cleaned_response)
        except json.JSONDecodeError:
            return None
    return None

def get_best_match_from_ai(item_edital, df_candidates, produto: dict, specs_config: dict):
    """Usa o modelo de IA para encontrar o melhor match dentro dos candidatos."""
    print(" - Asking AI for the best match...")

    produto_json = json.dumps(produto, ensure_ascii=False, indent=2)
    specs_config_json = json.dumps(specs_config, ensure_ascii=False, indent=2)
    candidates_json = df_candidates[['DESCRICAO','categoria_principal','subcategoria','MARCA','MODELO','VALOR']] \
                        .to_json(orient="records", force_ascii=False, indent=2)

    prompt = f"""<identidade>Você é um consultor de licitações com 20+ anos de experiência em áudio/instrumentos, focado na Lei 14.133/21, economicidade e menor preço.</identidade>
<objetivo>
1. Analise tecnicamente o produto do edital: {produto_json}
2. Use a estrutura de specs: {specs_config_json}. Considere 'required' como críticas, outros como opcionais.
3. Compare-o com cada produto na <base_fornecedores_filtrada>. Extraia specs de cada candidato se necessário.
4. **Seleção Primária**: Encontre o produto >=95% compatível (todas críticas + 100% opcionais). Dentre os compatíveis, escolha o de **menor 'Valor'**.
5. **Seleção Secundária**: Se nenhum, identifique o tecnicamente mais próximo e detalhe as especificações não atendidas, com contagens de críticas/opcionais atendidas.
6. Calcule score ponderado usando weights para críticas, e peso igual para opcionais.
7. Use o Google para pesquisar e confirmar as especificações técnicas dos produtos candidatos para garantir uma análise precisa.
8. Responda **apenas** com um objeto JSON seguindo o <formato_saida>.
</objetivo>

<formato_saida>
Responda APENAS com um único objeto JSON. Não inclua ```json, explicações ou qualquer outro texto.
**CASO 1: Encontrou um produto >=95% compatível.**
{{
  "best_match": {{
    "Marca": "Marca do Produto",
    "Modelo": "Modelo do Produto",
    "Valor": 1234.56,
    "Descricao_fornecedor": "Descrição completa do produto na base",
    "Compatibilidade_analise": "Análise quantitativa da compatibilidade. Ex: 'Compatível. Atende a 5 de 6 especificações chave.'",
    "criticas_atendidas": 3,
    "criticas_total": 3,
    "opcionais_atendidas": 4,
    "opcionais_total": 4,
    "score_ponderado": 98.5
  }}
}}

**CASO 2: NENHUM produto é >=95% compatível.**
Retorne "best_match" como `null`. Adicionalmente, inclua "closest_match" com os dados do produto mais próximo e "reasoning" explicando o motivo da incompatibilidade.
{{
  "best_match": null,
  "reasoning": "Explique o principal motivo da incompatibilidade. Ex: 'Nenhum produto atende à especificação de material X ou potência Y.'",
  "closest_match": {{
    "Marca": "Marca do Produto Mais Próximo",
    "Modelo": "Modelo do Mais Próximo",
    "Valor": 4321.98,
    "Descricao_fornecedor": "Descrição do produto mais próximo.",
    "Compatibilidade_analise": "Análise da compatibilidade parcial com detalhes. Ex: 'Atende [especificação A, B], mas falha em [especificação C (material), D (potência)].'",
    "criticas_atendidas": 2,
    "criticas_total": 3,
    "opcionais_atendidas": 3,
    "opcionais_total": 5,
    "score_ponderado": 75.0
  }}
}}
</formato_saida><base_fornecedores_filtrada>{candidates_json}</base_fornecedores_filtrada>"""

    response_text = gerar_conteudo_com_fallback(prompt, LLM_MODELS_FALLBACK)
    if response_text:
        MAX_RETRIES = 3
        for attempt in range(MAX_RETRIES):
            try:
                cleaned_response = response_text.strip().replace("```json", "").replace("```", "")
                return json.loads(cleaned_response)
            except json.JSONDecodeError as e:
                print(f"   - ERROR decoding JSON from AI (Attempt {attempt + 1}/{MAX_RETRIES}): {e}")
                if attempt < MAX_RETRIES - 1:
                    print("   - Retrying in 10 seconds...")
                    time.sleep(10)
                else:
                    print("   - Max retries reached. Returning error.")
                    return { "best_match": None, "closest_match": None, "reasoning": f"Erro na decodificação do JSON da API após {MAX_RETRIES} tentativas: {e}"}
    else:
        return { "best_match": None, "closest_match": None, "reasoning": "Falha na chamada da API para todos os modelos de fallback."}

def calculate_compatibility_score(ai_data: dict) -> float:
    """
    Calcula a pontuação de compatibilidade baseada nos dados da IA.
    """
    if 'score_ponderado' in ai_data:
        return ai_data['score_ponderado']
    return 0.0

def get_compatibility_level(score: float, criticas_atendidas: int, criticas_total: int, opcionais_atendidas: int, opcionais_total: int) -> str:
    if criticas_total == 0:
        return "INDEFINIDO"
    criticas_perc = criticas_atendidas / criticas_total
    if opcionais_total > 0:
        opcionais_perc = opcionais_atendidas / opcionais_total
    else:
        opcionais_perc = 1.0
    
    if score >= 95 and criticas_perc == 1:
        return "MATCH PERFEITO"
    elif score >= 85 and criticas_perc == 1 and opcionais_perc >= 0.8:
        return "MATCH EXCELENTE"
    elif score >= 70 and criticas_perc == 1 and opcionais_perc >= 0.6:
        return "MATCH BOM"
    elif score >= 50 and criticas_perc == 1:
        return "MATCH ACEITÁVEL"
    elif score >= 25 and criticas_perc > 0.5:
        return "MATCH PARCIAL"
    else:
        return "INCOMPATÍVEL"

def get_rainbow_color(score: float) -> PatternFill:
    """
    Gera cor com base no score de compatibilidade:
    - 95-100%: Verde escuro
    - 85-94%: Verde claro
    - 70-84%: Amarelo claro
    - 50-69%: Laranja
    - 25-49%: Vermelho
    - <25%: Vermelho escuro
    """
    if pd.isna(score):
        score = 0.0

    if score >= 95:
        hex_color = '006400'  # Verde escuro
    elif score >= 85:
        hex_color = '90EE90'  # Verde claro
    elif score >= 70:
        hex_color = 'FFFACD'  # Amarelo claro
    elif score >= 50:
        hex_color = 'FFA500'  # Laranja
    elif score >= 25:
        hex_color = 'FF0000'  # Vermelho
    else:
        hex_color = '8B0000'  # Vermelho escuro

    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

# ============================================================ 
# MAIN
# ============================================================ 

def main():
    logger.info("Starting the product matching process...")
    print("Starting the product matching process...")

    load_dotenv()
    api_key = os.getenv("GOOGLE_API_PAGO")
    if not api_key:
        logger.error("GOOGLE_API_KEY not found in .env file.")
        print("ERROR: GOOGLE_API_KEY not found in .env file.")
        return

    genai.configure(api_key=api_key)

    try:
        df_edital = pd.read_excel(CAMINHO_EDITAL)
        df_base = pd.read_excel(CAMINHO_BASE)
        
        # Convert 'VALOR' to numeric, coercing errors
        df_base['VALOR'] = pd.to_numeric(df_base['VALOR'], errors='coerce').fillna(0)

        logger.info(f"Loaded {len(df_edital)} items from {os.path.basename(CAMINHO_EDITAL)} and {len(df_base)} products from {os.path.basename(CAMINHO_BASE)}")
        print(f"👾 Edital loaded: {len(df_edital)} items from {os.path.basename(CAMINHO_EDITAL)}")
        print(f"📯 Product base loaded: {len(df_base)} products from {os.path.basename(CAMINHO_BASE)}")
    except FileNotFoundError as e:
        logger.error(f"Could not load data files. Details: {e}")
        print(f"ERROR: Could not load data files. Details: {e}")
        return

    if os.path.exists(CAMINHO_HEAVY_EXISTENTE):
        logger.info(f"Loading existing processed data from {os.path.basename(CAMINHO_HEAVY_EXISTENTE)}")
        print(f"   - Loading existing data from: {os.path.basename(CAMINHO_HEAVY_EXISTENTE)}")
        df_existing = pd.read_excel(CAMINHO_HEAVY_EXISTENTE)
        existing_keys = set(zip(df_existing['ARQUIVO'].astype(str), df_existing['Nº'].astype(str)))
        logger.info(f"Found {len(existing_keys)} already processed items.")
    else:
        df_existing = pd.DataFrame()
        existing_keys = set()
        logger.info("No existing output file found. Processing all items as new.")

    df_edital_new = df_edital[~df_edital.apply(lambda row: (str(row['ARQUIVO']), str(row['Nº'])) in existing_keys, axis=1)].copy()

    if df_edital_new.empty:
        logger.info("No new items to process. Output file is up to date.")
        print("\n✅ No new items to process. The output file is already up to date.")
        return

    logger.info(f"Identified {len(df_edital_new)} new items to process.")
    print(f"   - Found {len(df_edital_new)} new items to process.")
    total_new_items = len(df_edital_new)

    for idx, item_edital in df_edital_new.iterrows():
        descricao = str(item_edital['DESCRICAO'])
        referencia = str(item_edital.get('REFERENCIA', 'N/A'))
        print(f"\n 📈 Processing new item {idx + 1}/{total_new_items}: {descricao[:60]}...")
        status = ""
        best_match_data = None
        closest_match_data = None
        reasoning = None
        data_to_populate = None
        classification = None
        produto = None

        valor_unit_edital = float(str(item_edital.get('VALOR_UNIT', '0')).replace(',', '.'))

        if valor_unit_edital > 0:
            max_cost = valor_unit_edital * INITIAL_PRICE_FILTER_PERCENTAGE
            df_candidates = df_base[df_base['VALOR'] <= max_cost].copy()
        else:
            print("- Valor de referência do edital é R$0.00 ou inválido. Analisando todos os produtos da base.")
            df_candidates = df_base.copy()

        if df_candidates.empty:
            if valor_unit_edital > 0:
                print(f"- ⚠️ Nenhum produto encontrado abaixo do custo máximo de R${max_cost:.2f}.")
            status = "Nenhum Produto com Margem"
        else:
            classification = get_item_classification(descricao, referencia, CATEGORIZATION_KEYWORDS)
        
        time.sleep(10) 

        df_final_candidates = pd.DataFrame()
        filter_level = "Nenhum"
        specs_config = {}

        if classification:
            main_category = classification.get('categoria_principal')
            subcategory = classification.get('subcategoria')
            print(f"   - AI classified item as: Categoria='{main_category}', Subcategoria='{subcategory}'")
            specs_config = SPECS_BY_CATEGORY.get(subcategory.lower(), SPECS_BY_CATEGORY.get(main_category.lower(), {}))

            if subcategory:
                df_filtered_sub = df_candidates[df_candidates['subcategoria'].str.contains(subcategory, case=False, na=False)]
                if not df_filtered_sub.empty:
                    print(f"  - 📦 Found {len(df_filtered_sub)} candidates matching SUBCATEGORY '{subcategory}'.")
                    df_final_candidates = df_filtered_sub
                    filter_level = "Subcategoria"

            if df_final_candidates.empty and main_category:
                print(f"  - ⚠️ No candidates found for subcategory '{subcategory}'. Trying main category...")
                df_filtered_main = df_candidates[df_candidates['categoria_principal'] == main_category]
                if not df_filtered_main.empty:
                    print(f"  - 📦 Found {len(df_filtered_main)} candidates matching MAIN CATEGORY '{main_category}'.")
                    df_final_candidates = df_filtered_main
                    filter_level = "Categoria Principal"

            if df_final_candidates.empty:
                print(f"  - ⚠️ No candidates found for main category '{main_category}'. Using all price-filtered products...")
                df_final_candidates = df_candidates
                filter_level = "Apenas Preço"
        
        else:
            print("   - ⚠️ AI classification failed. Using all price-filtered products as fallback.")
            df_final_candidates = df_candidates
            filter_level = "Apenas Preço (Falha na IA)"

        if df_final_candidates.empty:
            print(f"- ⚠️ Nenhum produto encontrado após todas as tentativas de filtro.")
            status = "Nenhum Produto na Categoria"
        else:
            print(f" - Temos {len(df_final_candidates)} candidatos para a IA (filtrado por: {filter_level}).")
            produto = extract_edital_specs(descricao, referencia, specs_config)
            ai_result = get_best_match_from_ai(item_edital, df_final_candidates, produto, specs_config)
            time.sleep(30)

            best_match_data = ai_result.get("best_match")
            closest_match_data = ai_result.get("closest_match")
            reasoning = ai_result.get("reasoning")

            if best_match_data:
                print(f" ✅ - AI recomenda: {best_match_data.get('Marca', 'N/A')} {best_match_data.get('Modelo', 'N/A')}")
                status = "Match Encontrado"
                data_to_populate = best_match_data
            elif closest_match_data:
                print(f"- 🧿 AI sugere como mais próximo: {closest_match_data.get('Marca', 'N/A')} {closest_match_data.get('Modelo', 'N/A')}")
                if reasoning:
                    print(f"    Motivo: {reasoning}")
                status = "Match Parcial (Sugestão)"
                data_to_populate = closest_match_data
            else:
                print(" ❌ - AI não encontrou produto compatível ou próximo.")
                if reasoning:
                    print(f"    Motivo: {reasoning}")
                status = "Nenhum Produto Compatível"

        result_row = {
            'ARQUIVO': item_edital['ARQUIVO'],
            'Nº': item_edital['Nº'],
            'DESCRICAO': item_edital['DESCRICAO'],
            'REFERENCIA': item_edital.get('REFERENCIA'),
            'UNID_FORN': item_edital.get('UNID_FORN'),
            'QTDE': item_edital.get('QTDE'),
            'VALOR_TOTAL': item_edital.get('VALOR_TOTAL'),
            'LOCAL_ENTREGA': item_edital.get('LOCAL_ENTREGA'),
            'INTERVALO_LANCES': item_edital.get('INTERVALO_LANCES'),
            'VALOR_UNIT_EDITAL': item_edital['VALOR_UNIT'],
            'STATUS': status,
            'MOTIVO_INCOMPATIBILIDADE': reasoning if status != "Match Encontrado" else None,
            'LAST_UPDATE': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        # Popula a coluna PRODUTO com o estereótipo gerado pela IA
        if produto:
            categoria = produto.get('categoria_primaria', 'N/A')
            atributos = produto.get('atributos', {})
            atributos_str = ', '.join([f'{k}: {v}' for k, v in atributos.items()])
            stereotype_str = f"Categoria: {categoria} | Atributos: {atributos_str}"
            result_row['PRODUTO'] = stereotype_str
        else:
            result_row['PRODUTO'] = "Falha ao gerar estereótipo"

        if data_to_populate:
            cost_price = float(data_to_populate.get('Valor') or 0)
            final_price = cost_price * (1 + PROFIT_MARGIN)
            margem_lucro_valor = final_price - cost_price
            qtde_val = item_edital.get('QTDE')
            qtde = 0
            if pd.notna(qtde_val):
                try:
                    qtde = int(float(qtde_val))
                except (ValueError, TypeError):
                    qtde = 0
            lucro_total = margem_lucro_valor * qtde
            # --- Nova lógica de compatibilidade ---
            analise_compat = data_to_populate.get('Compatibilidade_analise')
            
            # Obter subcategoria do item do edital
            edital_subcategory = classification.get('subcategoria') if classification else None

            # Obter subcategoria do produto sugerido
            product_subcategory = None
            desc_fornecedor = data_to_populate.get('Descricao_fornecedor')
            if desc_fornecedor:
                # Usar df_final_candidates que foi a base para a IA
                matched_rows = df_final_candidates[df_final_candidates['DESCRICAO'] == desc_fornecedor]
                if not matched_rows.empty:
                    product_subcategory = matched_rows.iloc[0]['subcategoria']

            criticas_atendidas = data_to_populate.get('criticas_atendidas', 0)
            criticas_total = data_to_populate.get('criticas_total', 0)
            opcionais_atendidas = data_to_populate.get('opcionais_atendidas', 0)
            opcionais_total = data_to_populate.get('opcionais_total', 0)
            compat_score = calculate_compatibility_score(data_to_populate)
            nivel_compat = get_compatibility_level(compat_score, criticas_atendidas, criticas_total, opcionais_atendidas, opcionais_total)
            # A linha que criava 'produto_str' foi removida pois não é mais necessária para a coluna 'PRODUTO'

            result_row.update({
                'MARCA_SUGERIDA': data_to_populate.get('Marca'),
                'MODELO_SUGERIDO': data_to_populate.get('Modelo'),
                'CUSTO_FORNECEDOR': cost_price,
                'PRECO_FINAL_VENDA': final_price,
                'MARGEM_LUCRO_VALOR': margem_lucro_valor,
                'LUCRO_TOTAL': lucro_total,
                'DESCRICAO_FORNECEDOR': desc_fornecedor,
                'ANALISE_COMPATIBILIDADE': analise_compat,
                'COMPATIBILITY_SCORE': compat_score,
                'NIVEL_COMPATIBILIDADE': nivel_compat
            })

        # Append the result to existing data
        df_existing = pd.concat([df_existing, pd.DataFrame([result_row])], ignore_index=True)

        # Save incrementally
        df_final = df_existing

        output_columns = [
            'ARQUIVO','Nº','DESCRICAO','REFERENCIA','PRODUTO','STATUS',
            'UNID_FORN', 'QTDE', 'VALOR_UNIT_EDITAL', 'VALOR_TOTAL',
            'LOCAL_ENTREGA', 'INTERVALO_LANCES',
            'MARCA_SUGERIDA', 'MODELO_SUGERIDO', 'CUSTO_FORNECEDOR',
            'PRECO_FINAL_VENDA','MARGEM_LUCRO_VALOR', 'LUCRO_TOTAL', 'MOTIVO_INCOMPATIBILIDADE',
            'DESCRICAO_FORNECEDOR','ANALISE_COMPATIBILIDADE','COMPATIBILITY_SCORE','NIVEL_COMPATIBILIDADE','LAST_UPDATE'
        ]
        for col in output_columns:
            if col not in df_final.columns:
                df_final[col] = ''
        df_final = df_final[output_columns]

        output_dir = os.path.dirname(CAMINHO_SAIDA)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        writer = pd.ExcelWriter(CAMINHO_SAIDA, engine='openpyxl')
        df_final.to_excel(writer, index=False, sheet_name='Proposta')
        workbook = writer.book
        worksheet = writer.sheets['Proposta']

        compat_col_idx = output_columns.index('COMPATIBILITY_SCORE') + 1  # 1-based

        for row_idx in range(2, len(df_final) + 2):
            score = df_final.at[row_idx - 2, 'COMPATIBILITY_SCORE']
            if pd.isna(score):
                score = 0.0
            color_fill = get_rainbow_color(score)
            for col_idx in range(1, len(output_columns) + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = color_fill

        writer.close()
        logger.info(f"Incremental save after processing item {idx + 1}/{total_new_items} at {CAMINHO_SAIDA}")
        print(f"Incremental save completed for item {idx + 1}/{total_new_items}.")

    logger.info("All new items processed and saved incrementally.")
    print("✅ All new items processed and saved incrementally.")

if __name__ == "__main__":
    main()