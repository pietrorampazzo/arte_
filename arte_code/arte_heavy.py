import pandas as pd
import google.generativeai as genai
import google.api_core.exceptions as google_exceptions
import os
from dotenv import load_dotenv
import json
import time
import logging
from openpyxl.styles import PatternFill
import re
from datetime import datetime
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np

# ======================================================================
# CONFIGURAÇÕES E CONSTANTES
# ======================================================================

# --- File Paths ---
BASE_DIR = r"C:\Users\pietr\OneDrive\.vscode\arte_"
CAMINHO_EDITAL = os.path.join(BASE_DIR, "DOWNLOADS", "master.xlsx")
CAMINHO_BASE = os.path.join(BASE_DIR, "DOWNLOADS", "METADADOS", "produtos_metadados.xlsx")
CAMINHO_SAIDA = os.path.join(BASE_DIR, "DOWNLOADS", "master_heavy.xlsx")
CAMINHO_HEAVY_EXISTENTE = CAMINHO_SAIDA

# --- Financial Parameters ---
PROFIT_MARGIN = 0.53  # MARGEM DE LUCRO
INITIAL_PRICE_FILTER_PERCENTAGE = 0.60  # FILTRO DE PREÇO INICIAL
EXPANDED_PRICE_FILTER_PERCENTAGE = 0.75 # FILTRO DE PREÇO EXPANDIDO

# --- AI Model Configuration ---
LLM_MODELS_FALLBACK = [
    "gemini-2.5-pro",
    "gemini-2.5-flash",
    "gemini-2.5-flash-lite",
    "gemini-2.0-flash-lite",
    "gemini-2.0-flash",
    "gemini-1.5-flash",  
]

# --- ML Configuration ---
SUBCATEGORY_ML_CANDIDATES = 10
MAIN_CATEGORY_ML_CANDIDATES = 20

# --- Categorization Keywords ---
CATEGORIZATION_KEYWORDS = {
        "EQUIPAMENTO_SOM" : ["caixa_ativa", "caixa_passiva", "caixa_portatil", "line_array", "subwoofer_ativo", "subwoofer_passivo", "amplificador_potencia", "cabeçote_amplificado", "coluna_vertical", "monitor_de_palco"],
        "EQUIPAMENTO_AUDIO" : ["microfone_dinamico", "microfone_condensador", "microfone_lapela", "microfone_sem_fio", "microfone_instrumento", "mesa_analogica", "mesa_digital", "interface_audio", "processador_dsp", "fone_monitor", "sistema_iem", "pedal_efeitos"],
        "INSTRUMENTO_CORDA" : ["violao", "guitarra", "contra_baixo", "violino", "violoncelo", "ukulele", "cavaquinho"],
        "INSTRUMENTO_PERCUSSAO" : ["bateria_acustica", "bateria_eletronica", "repinique", "rocari", "tantan", "rebolo","surdo_mao", "cuica", "zabumba", "caixa_guerra", "bombo_fanfarra", "lira_marcha","tarol", "malacacheta", "caixa_bateria", "pandeiro", "tamborim","reco_reco", "agogô", "triangulo", "chocalho", "afuche", "cajon", "bongo", "conga", "djembé", "timbal", "atabaque", "berimbau","tam_tam", "caxixi", "carilhao", "xequerê", "prato"],
        "INSTRUMENTO_SOPRO" : ["saxofone", "trompete", "trombone", "trompa", "clarinete", "flauta", "tuba", "flugelhorn", "bombardino", "corneta", "cornetão"],
        "INSTRUMENTO_TECLADO" : ["teclado_digital", "piano_acustico", "piano_digital", "sintetizador", "controlador_midi", "glockenspiel", "metalofone"],
        "ACESSORIO_MUSICAL" : ["banco_teclado", "estante_partitura", "suporte_microfone", "suporte_instrumento", "carrinho_transporte", "case_bag", "afinador", "metronomo", "cabos_audio", "palheta", "cordas", "oleo_lubrificante", "graxa", "surdina", "bocal_trompete", "pele_percussao", "baqueta", "talabarte", "pedal_bumbo", "chimbal_hihat"],
        "EQUIPAMENTO_TECNICO" : ["ssd", "fonte_energia", "switch_rede", "projetor", "drone"]
}

# --- Logging Configuration ---
LOG_FILE = os.path.join(BASE_DIR, "LOGS", "arte_stanley.log")
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ============================================================
# FUNÇÕES DE IA E ML
# ============================================================

def gerar_conteudo_com_fallback(prompt: str, modelos: list[str]) -> str | None:
    """Tenta gerar conteúdo usando uma lista de modelos em ordem de preferência."""
    for nome_modelo in modelos:
        try:
            print(f"   - Tentando chamada à API com o modelo: {nome_modelo}...")
            model = genai.GenerativeModel(nome_modelo)
            response = model.generate_content(prompt)

            if not response.parts:
                finish_reason = response.candidates[0].finish_reason.name if response.candidates else 'N/A'
                print(f"   - ❌ A GERAÇÃO RETORNOU VAZIA. Motivo: {finish_reason}.")
                return None
            print(f"   - Sucesso com o modelo '{nome_modelo}'.")
            return response.text
        except google_exceptions.ResourceExhausted as e:
            print(f"- Cota excedida para o modelo '{nome_modelo}'. Tentando o próximo da lista.")
            time.sleep(5)
            continue
        except Exception as e:
            print(f"   - ❌ Erro inesperado com o modelo '{nome_modelo}': {e}")
            return None
    print("   - ❌ FALHA TOTAL: Todos os modelos na lista de fallback falharam.")
    return None

def get_item_classification(description: str, reference: str, categories_with_subcategories: dict) -> dict | None:
    """Usa o modelo de IA para classificar o item."""
    print("- Asking AI for item classification (category and subcategory)...")
    prompt = f"""<identidade>Você é um especialista de almoxarifado e perito em catalogação de produtos.</identidade>
<objetivo>
Sua tarefa é classificar o item a seguir, identificando sua `categoria_principal` e `subcategoria` com base na estrutura fornecida.
- A `categoria_principal` DEVE ser uma das chaves da estrutura.
- A `subcategoria` DEVE ser um dos valores da lista associada à categoria principal escolhida.
Responda APENAS com um objeto JSON.
</objetivo>
<item_a_ser_classificado>
Descrição: {description}
Referência: {reference}
</item_a_ser_classificado>
<estrutura_de_categorias_e_subcategorias_permitidas>
{json.dumps(categories_with_subcategories, indent=2, ensure_ascii=False)}
</estrutura_de_categorias_e_subcategorias_permitidas>
<formato_saida>
{{
  "categoria_principal": "NOME_DA_CATEGORIA_PRINCIPAL",
  "subcategoria": "NOME_DA_SUBCATEGORIA_ESPECIFICA"
}}
</formato_saida>
JSON:
"""
    response_text = gerar_conteudo_com_fallback(prompt, LLM_MODELS_FALLBACK)
    if response_text:
        try:
            cleaned_response = response_text.strip().replace("```json", "").replace("```", "")
            classification = json.loads(cleaned_response)
            if 'categoria_principal' in classification and 'subcategoria' in classification and classification['categoria_principal'] in categories_with_subcategories:
                return classification
            else:
                print(f"   - WARNING: AI returned an invalid or incomplete classification: {classification}")
                return None
        except json.JSONDecodeError as e:
            print(f"   - ERROR decoding JSON from AI for classification: {e}")
            return None
    return None

def get_best_match_from_ai(item_edital, df_candidates, attempt_description: str):
    """Usa o modelo de IA para encontrar o melhor match dentro dos candidatos, retornando uma análise estruturada."""
    print(f" - Asking AI for the best match ({attempt_description})...")
    if df_candidates.empty:
        print("   - WARNING: Candidate DataFrame is empty. Skipping AI call.")
        return {"best_match": None, "closest_match": None, "reasoning": "No candidates provided."}

    candidates_json = df_candidates[['DESCRICAO','categoria_principal','subcategoria','MARCA','MODELO','VALOR']].to_json(orient="records", force_ascii=False, indent=2)
    prompt = f"""<identidade>Você é um consultor de licitações com 20+ anos de experiência em áudio/instrumentos, focado na Lei 14.133/21, economicidade e menor preço.</identidade>
<objetivo>
1.  Analise tecnicamente o item do edital: Descrição: `"{item_edital['DESCRICAO']}"` Referência: `"{item_edital.get('REFERENCIA', 'N/A')}"`.
2.  Compare-o com cada produto na `<base_fornecedores_filtrada>`.
3.  **Seleção Primária**: Encontre o produto da base que seja >=95% compatível. Dentre os compatíveis, escolha o de **menor 'Valor'**.
4.  **Seleção Secundária**: Se nenhum for >=95% compatível, identifique o produto tecnicamente mais próximo.
5.  **Análise Estruturada**: Para o produto escolhido (seja `best_match` ou `closest_match`), forneça uma análise de compatibilidade detalhada no formato JSON especificado.
6.  **Falhas Críticas**: Se uma especificação chave e obrigatória do edital (ex: sensor térmico, potência mínima, marca específica) não for atendida, adicione a flag `FALHA CRÍTICA:` no início do ponto negativo correspondente.
7.  Responda **apenas** com um objeto JSON.
</objetivo>
<formato_saida>
Responda APENAS com um único objeto JSON. Não inclua ```json ou qualquer outro texto.

**CASO 1: Encontrou um produto >=95% compatível.**
{{
  "best_match": {{
    "Marca": "Marca do Produto",
    "Modelo": "Modelo do Produto",
    "Valor": 1234.56,
    "Descricao_fornecedor": "Descrição completa do produto na base",
    "Compatibilidade_analise": {{
      "score_proposto": 98,
      "pontos_positivos": ["Atende à potência de 50W", "Material é alumínio conforme solicitado"],
      "pontos_negativos": ["Não inclui o cabo de força (acessório menor)"],
      "justificativa": "Excelente compatibilidade, apenas uma pequena divergência de acessório que não compromete o uso principal."
    }}
  }}
}}

**CASO 2: NENHUM produto é >=95% compatível.**
Retorne `best_match` como `null` e preencha `closest_match` com a análise do produto mais próximo.
{{
  "best_match": null,
  "closest_match": {{
    "Marca": "Marca do Produto Mais Próximo",
    "Modelo": "Modelo do Mais Próximo",
    "Valor": 4321.98,
    "Descricao_fornecedor": "Descrição do produto mais próximo.",
    "Compatibilidade_analise": {{
      "score_proposto": 40,
      "pontos_positivos": ["É da mesma marca", "Função de voo está presente"],
      "pontos_negativos": ["FALHA CRÍTICA: Não possui sensor térmico", "FALHA CRÍTICA: Zoom de 2x ao invés de 56x", "Não possui controle com tela"],
      "justificativa": "Incompatível para o uso pretendido devido à ausência de funcionalidades críticas (sensor térmico e zoom)."
    }}
  }}
}}
</formato_saida><base_fornecedores_filtrada>{candidates_json}</base_fornecedores_filtrada>"""

    response_text = gerar_conteudo_com_fallback(prompt, LLM_MODELS_FALLBACK)
    if response_text:
        try:
            cleaned_response = response_text.strip().replace("```json", "").replace("```", "")
            return json.loads(cleaned_response)
        except json.JSONDecodeError as e:
            print(f"   - ERROR decoding JSON from AI (Attempt: {attempt_description}): {e}")
            return {"best_match": None, "closest_match": None, "reasoning": f"Erro na decodificação do JSON da API: {e}"}
    return {"best_match": None, "closest_match": None, "reasoning": "Falha na chamada da API para todos os modelos de fallback."}

def get_top_n_ml_matches(item_edital, df_candidates, n):
    """Usa TF-IDF e Cosine Similarity para encontrar os N produtos mais similares."""
    if df_candidates.empty or n == 0:
        return pd.DataFrame()

    print(f" - Running ML to find top {n} matches...")
    edital_text = f"{item_edital['DESCRICAO']} {item_edital.get('REFERENCIA', '')}"
    candidates_texts = df_candidates['DESCRICAO'].fillna('').tolist()

    if not candidates_texts:
        return pd.DataFrame()

    vectorizer = TfidfVectorizer()
    all_texts = [edital_text] + candidates_texts
    tfidf_matrix = vectorizer.fit_transform(all_texts)

    edital_vec = tfidf_matrix[0:1]
    candidates_vec = tfidf_matrix[1:]

    similarities = cosine_similarity(edital_vec, candidates_vec).flatten()

    # Get indices of top N similarities
    num_results = min(n, len(similarities))
    top_indices = np.argsort(similarities)[-num_results:][::-1]

    print(f"   - Found {len(top_indices)} ML candidates.")
    return df_candidates.iloc[top_indices]

# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def check_match_found(ai_result: dict) -> bool:
    """Verifica se o resultado da IA contém um 'best_match' válido."""
    best_match_data = ai_result.get("best_match")
    if not best_match_data or not isinstance(best_match_data, dict):
        return False

    analise = best_match_data.get('Compatibilidade_analise')
    if not isinstance(analise, dict):
        return False # A análise agora deve ser um objeto

    # Um 'best_match' é considerado encontrado se a IA o retornou no campo correto.
    # A qualidade desse match será refletida no score.
    return True

def calculate_compatibility_score(analise_obj: dict | str) -> float:
    """
    Calcula a pontuação de compatibilidade (0-100) de forma estruturada,
    baseando-se no objeto JSON retornado pela IA.
    """
    # Fallback para formato antigo ou erros, garantindo que não quebre o processo.
    if not isinstance(analise_obj, dict):
        return 0.0

    justificativa = analise_obj.get('justificativa', '')
    pontos_positivos = analise_obj.get('pontos_positivos', [])
    pontos_negativos = analise_obj.get('pontos_negativos', [])
    score_proposto = analise_obj.get('score_proposto')

    # Regra 1: Presença de 'FALHA CRÍTICA' reduz drasticamente o score.
    if any('FALHA CRÍTICA' in str(p).upper() for p in pontos_negativos):
        return 20.0  # Score baixo para indicar inviabilidade, mas não é zero.

    # Regra 2: Compatibilidade total e explícita.
    if not pontos_negativos:
        if isinstance(score_proposto, (int, float)) and score_proposto >= 99:
            return 100.0
        if 'totalmente compatível' in justificativa.lower() or 'atende a todas' in justificativa.lower():
            return 100.0

    # Regra 3: Usar o score proposto pela IA como fonte principal.
    if isinstance(score_proposto, (int, float)) and 0 <= score_proposto <= 100:
        # Ajuste: Se a IA der 100, mas houver pontos negativos menores, limitamos a 95.
        if score_proposto == 100 and pontos_negativos:
            return 95.0
        return float(score_proposto)

    # Regra 4: Fallback quantitativo se o score da IA não for confiável.
    total_pontos = len(pontos_positivos) + len(pontos_negativos)
    if total_pontos > 0:
        score = (len(pontos_positivos) / total_pontos) * 100
        # Se o cálculo resultar em 100, mas houver pontos negativos, é uma contradição. Limitar a 95.
        if score == 100 and pontos_negativos:
            return 95.0
        return score

    # Regra 5: Se a análise for genericamente positiva, mas sem detalhes, usar 85 para revisão.
    if 'compatível' in justificativa.lower() and not pontos_negativos and not pontos_positivos:
        return 85.0

    return 0.0

def get_rainbow_color(score: float) -> PatternFill:
    """Gera cor com base no score de compatibilidade."""
    if pd.isna(score):
        score = 0.0
    strong_green, light_green, yellow, red = ('00B050', 'C6EFCE', 'FFEB9C', 'FFC7CE')
    if score == 100: hex_color = strong_green
    elif score >= 80: hex_color = light_green
    elif score >= 40: hex_color = yellow
    else: hex_color = red
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def process_single_item_pipeline(item_edital, df_base, price_filter_percentage, classification):
    """Executa o pipeline de análise para um único item do edital."""
    descricao = str(item_edital['DESCRICAO'])
    referencia = str(item_edital.get('REFERENCIA', 'N/A'))
    valor_unit_edital = float(str(item_edital.get('VALOR_UNIT', '0')).replace(',', '.'))
    
    # Filtro de Preço
    if valor_unit_edital > 0:
        max_cost = valor_unit_edital * price_filter_percentage
        df_price_candidates = df_base[df_base['VALOR'] <= max_cost].copy()
    else:
        print("- Valor de referência do edital é R$0.00 ou inválido. Analisando todos os produtos da base.")
        df_price_candidates = df_base.copy()

    if df_price_candidates.empty:
        return "Nenhum Produto com Margem", None, None

    if not classification:
        print("   - ⚠️ AI classification failed. Cannot proceed with category filters.")
        # Fallback: usar ML em todos os produtos filtrados por preço
        df_ml_candidates = get_top_n_ml_matches(item_edital, df_price_candidates, MAIN_CATEGORY_ML_CANDIDATES)
        ai_result = get_best_match_from_ai(item_edital, df_ml_candidates, "Fallback ML on all price-filtered")
        time.sleep(20)
        best_match_data = ai_result.get("best_match")
        closest_match_data = ai_result.get("closest_match")
        if best_match_data:
            print(f"   ✅ - AI recomenda: {best_match_data.get('Marca', 'N/A')} {best_match_data.get('Modelo', 'N/A')}")
        elif closest_match_data:
            print(f"   - 🧿 AI sugere como mais próximo: {closest_match_data.get('Marca', 'N/A')} {closest_match_data.get('Modelo', 'N/A')}")

        if check_match_found(ai_result):
            return "Match Encontrado (Fallback ML)", ai_result.get("best_match"), None
        else:
            return "Nenhum Produto Compatível", ai_result.get("best_match"), ai_result.get("closest_match")

    main_category = classification.get('categoria_principal')
    subcategory = classification.get('subcategoria')
    print(f"   - AI classified item as: Categoria='{main_category}', Subcategoria='{subcategory}'")

    # --- Etapa 1: LLM na Subcategoria ---
    print("\n--- ETAPA 1: LLM na Subcategoria ---")
    df_filtered_sub = df_price_candidates[df_price_candidates['subcategoria'].str.contains(subcategory, case=False, na=False)]
    if not df_filtered_sub.empty:
        print(f"  - 📦 Found {len(df_filtered_sub)} candidates matching SUBCATEGORY '{subcategory}'.")
        ai_result = get_best_match_from_ai(item_edital, df_filtered_sub, "Subcategory Filter")
        time.sleep(20)
        best_match_data = ai_result.get("best_match")
        closest_match_data = ai_result.get("closest_match")
        if best_match_data:
            print(f"   ✅ - AI recomenda: {best_match_data.get('Marca', 'N/A')} {best_match_data.get('Modelo', 'N/A')}")
        elif closest_match_data:
            print(f"   - 🧿 AI sugere como mais próximo: {closest_match_data.get('Marca', 'N/A')} {closest_match_data.get('Modelo', 'N/A')}")

        if check_match_found(ai_result):
            return "Match Encontrado (Subcategoria)", ai_result.get("best_match"), None
    else:
        print(f"  - ⚠️ No candidates found for subcategory '{subcategory}'.")

    # --- Etapa 2: ML na Subcategoria + LLM ---
    print("\n--- ETAPA 2: ML na Subcategoria + LLM ---")
    if not df_filtered_sub.empty:
        df_ml_sub_candidates = get_top_n_ml_matches(item_edital, df_filtered_sub, SUBCATEGORY_ML_CANDIDATES)
        ai_result = get_best_match_from_ai(item_edital, df_ml_sub_candidates, f"ML Top {SUBCATEGORY_ML_CANDIDATES} in Subcategory")
        time.sleep(20)
        best_match_data = ai_result.get("best_match")
        closest_match_data = ai_result.get("closest_match")
        if best_match_data:
            print(f"   ✅ - AI recomenda: {best_match_data.get('Marca', 'N/A')} {best_match_data.get('Modelo', 'N/A')}")
        elif closest_match_data:
            print(f"   - 🧿 AI sugere como mais próximo: {closest_match_data.get('Marca', 'N/A')} {closest_match_data.get('Modelo', 'N/A')}")

        if check_match_found(ai_result):
            return f"Match Encontrado (ML na Subcategoria)", ai_result.get("best_match"), None

    # --- Etapa 3: ML na Categoria Principal + LLM ---
    print("\n--- ETAPA 3: ML na Categoria Principal + LLM ---")
    df_filtered_main = df_price_candidates[df_price_candidates['categoria_principal'] == main_category]
    if not df_filtered_main.empty:
        print(f"  - 📦 Found {len(df_filtered_main)} candidates matching MAIN CATEGORY '{main_category}'.")
        df_ml_main_candidates = get_top_n_ml_matches(item_edital, df_filtered_main, MAIN_CATEGORY_ML_CANDIDATES)
        ai_result = get_best_match_from_ai(item_edital, df_ml_main_candidates, f"ML Top {MAIN_CATEGORY_ML_CANDIDATES} in Main Category")
        time.sleep(20)
        best_match_data = ai_result.get("best_match")
        closest_match_data = ai_result.get("closest_match")
        if best_match_data:
            print(f"   ✅ - AI recomenda: {best_match_data.get('Marca', 'N/A')} {best_match_data.get('Modelo', 'N/A')}")
        elif closest_match_data:
            print(f"   - 🧿 AI sugere como mais próximo: {closest_match_data.get('Marca', 'N/A')} {closest_match_data.get('Modelo', 'N/A')}")

        if check_match_found(ai_result):
            return f"Match Encontrado (ML na Categoria Principal)", ai_result.get("best_match"), None
        else:
            # Se chegamos até aqui, este é o nosso melhor palpite
            return "Nenhum Match >95%", ai_result.get("best_match"), ai_result.get("closest_match")

    return "Nenhum Produto na Categoria", None, None

# ============================================================
# MAIN
# ============================================================

def main():
    logger.info("Starting the Stanley product matching process...")
    print("Starting the Stanley product matching process...")

    load_dotenv()
    api_key = os.getenv("GOOGLE_API_PAGO")
    if not api_key:
        logger.error("GOOGLE_API_KEY not found in .env file.")
        return
    genai.configure(api_key=api_key)

    try:
        df_edital = pd.read_excel(CAMINHO_EDITAL)
        df_base = pd.read_excel(CAMINHO_BASE)
        df_base['VALOR'] = pd.to_numeric(df_base['VALOR'], errors='coerce').fillna(0)
        logger.info(f"Loaded {len(df_edital)} items from edital and {len(df_base)} products from base.")
        print(f"👾 Edital loaded: {len(df_edital)} items.")
        print(f"📯 Product base loaded: {len(df_base)} products.")
    except FileNotFoundError as e:
        logger.error(f"Could not load data files. Details: {e}")
        return

    if os.path.exists(CAMINHO_HEAVY_EXISTENTE):
        logger.info(f"Loading existing processed data from {os.path.basename(CAMINHO_HEAVY_EXISTENTE)}")
        df_existing = pd.read_excel(CAMINHO_HEAVY_EXISTENTE)
        # Garantir consistência na chave de verificação, tratando 'Nº' como inteiro antes de string
        df_existing['ARQUIVO_str'] = df_existing['ARQUIVO'].astype(str).str.strip()
        df_existing['N_int_str'] = pd.to_numeric(df_existing['Nº'], errors='coerce').fillna(0).astype(int).astype(str)
        existing_keys = set(zip(df_existing['ARQUIVO_str'], df_existing['N_int_str']))
    else:
        df_existing = pd.DataFrame()
        existing_keys = set()

    # Criar chaves consistentes para o DataFrame do edital também
    df_edital['ARQUIVO_str'] = df_edital['ARQUIVO'].astype(str).str.strip()
    df_edital['N_int_str'] = pd.to_numeric(df_edital['Nº'], errors='coerce').fillna(0).astype(int).astype(str)

    # Filtrar itens novos usando as chaves normalizadas
    df_edital_new = df_edital[~df_edital.apply(lambda row: (row['ARQUIVO_str'], row['N_int_str']) in existing_keys, axis=1)].copy()

    if df_edital_new.empty:
        logger.info("No new items to process. Output file is up to date.")
        print("\n✅ No new items to process. The output file is already up to date.")
        return

    logger.info(f"Identified {len(df_edital_new)} new items to process.")
    print(f"   - Found {len(df_edital_new)} new items to process.")
    total_new_items = len(df_edital_new)

    for idx, item_edital in df_edital_new.iterrows():
        item_index_in_df = df_edital_new.index.get_loc(idx)
        print(f"\n📈 Processing new item {item_index_in_df + 1}/{total_new_items}: {str(item_edital['DESCRICAO'])[:60]}...")

        # --- ETAPA DE CLASSIFICAÇÃO (FEITA APENAS UMA VEZ) ---
        classification = get_item_classification(
            str(item_edital['DESCRICAO']),
            str(item_edital.get('REFERENCIA', 'N/A')),
            CATEGORIZATION_KEYWORDS
        )
        time.sleep(10)

        # --- ETAPA PADRÃO ---
        print("\n===== TENTATIVA 1: Filtro de Preço Padrão (60%) =====")
        status, best_match_data, closest_match_data = process_single_item_pipeline(
            item_edital, df_base, INITIAL_PRICE_FILTER_PERCENTAGE, classification
        )

        # --- ETAPA 4: Aumentar filtro de preço e repetir ---
        if "Match Encontrado" not in status:
            print("\n===== TENTATIVA 2: Filtro de Preço Expandido (75%) =====")
            logger.warning(f"Item {item_edital['Nº']} não encontrou match. Tentando com filtro de preço expandido.")
            status_exp, best_match_data_exp, closest_match_data_exp = process_single_item_pipeline(
                item_edital, df_base, EXPANDED_PRICE_FILTER_PERCENTAGE, classification
            )
            # Prioriza o resultado da tentativa expandida se encontrar um match
            if "Match Encontrado" in status_exp:
                status, best_match_data, closest_match_data = status_exp, best_match_data_exp, closest_match_data_exp
            # Se a tentativa expandida também não achou, mas tem uma sugestão melhor, usa ela
            elif closest_match_data_exp and not closest_match_data:
                 status, best_match_data, closest_match_data = status_exp, best_match_data_exp, closest_match_data_exp

        # Determina os dados finais para popular a linha
        data_to_populate = best_match_data if best_match_data else closest_match_data
        reasoning = None
        if not best_match_data and closest_match_data:
            status = "Match Parcial (Sugestão)"
            # O reasoning já vem da chamada da API

        result_row = {
            'ARQUIVO': item_edital['ARQUIVO'],
            'Nº': item_edital['Nº'],
            'DESCRICAO_EDITAL': item_edital['DESCRICAO'],
            'REFERENCIA': item_edital.get('REFERENCIA'),
            'UNID_FORN': item_edital.get('UNID_FORN'),
            'QTDE': item_edital.get('QTDE'),
            'VALOR_TOTAL': item_edital.get('VALOR_TOTAL'),
            'LOCAL_ENTREGA': item_edital.get('LOCAL_ENTREGA'),
            'INTERVALO_LANCES': item_edital.get('INTERVALO_LANCES'),
            'VALOR_UNIT_EDITAL': item_edital['VALOR_UNIT'],
            'STATUS': status,
            'MOTIVO_INCOMPATIBILIDADE': reasoning,
            'LAST_UPDATE': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        if data_to_populate:
            cost_price = float(data_to_populate.get('Valor') or 0)
            final_price = cost_price * (1 + PROFIT_MARGIN)
            margem_lucro_valor = final_price - cost_price
            qtde = 0
            qtde_val = item_edital.get('QTDE')
            if pd.notna(qtde_val):
                try:
                    qtde = int(float(qtde_val))
                except (ValueError, TypeError):
                    qtde = 0
            lucro_total = margem_lucro_valor * qtde

            analise_compat_obj = data_to_populate.get('Compatibilidade_analise')
            compat_score = calculate_compatibility_score(analise_compat_obj)

            # Gera uma descrição textual informativa a partir do objeto de análise
            analise_compat_text = ""
            if isinstance(analise_compat_obj, dict):
                justificativa = analise_compat_obj.get('justificativa', 'Análise não fornecida.')
                positivos = analise_compat_obj.get('pontos_positivos', [])
                negativos = analise_compat_obj.get('pontos_negativos', [])
                
                texto_prós = "Prós: " + "; ".join(positivos) if positivos else ""
                texto_contras = "Contras: " + "; ".join(negativos) if negativos else ""
                
                analise_compat_text = f"Justificativa: {justificativa}"
                if texto_prós:
                    analise_compat_text += f" | {texto_prós}"
                if texto_contras:
                    analise_compat_text += f" | {texto_contras}"
            elif analise_compat_obj: # Fallback para o formato antigo de string
                analise_compat_text = str(analise_compat_obj)

            result_row.update({
                'MARCA_SUGERIDA': data_to_populate.get('Marca'),
                'MODELO_SUGERIDO': data_to_populate.get('Modelo'),
                'CUSTO_FORNECEDOR': cost_price,
                'PRECO_FINAL_VENDA': final_price,
                'MARGEM_LUCRO_VALOR': margem_lucro_valor,
                'LUCRO_TOTAL': lucro_total,
                'DESCRICAO_FORNECEDOR': data_to_populate.get('Descricao_fornecedor'),
                'ANALISE_COMPATIBILIDADE': analise_compat_text, # Usa o novo texto formatado
                'COMPATIBILITY_SCORE': compat_score
            })

        df_existing = pd.concat([df_existing, pd.DataFrame([result_row])], ignore_index=True)

        # Save incrementally
        output_columns = [
            'ARQUIVO','Nº','DESCRICAO_EDITAL','REFERENCIA','STATUS',
            'UNID_FORN', 'QTDE', 'VALOR_UNIT_EDITAL', 'VALOR_TOTAL',
            'LOCAL_ENTREGA', 'INTERVALO_LANCES',
            'MARCA_SUGERIDA', 'MODELO_SUGERIDO', 'CUSTO_FORNECEDOR',
            'PRECO_FINAL_VENDA','MARGEM_LUCRO_VALOR', 'LUCRO_TOTAL', 'MOTIVO_INCOMPATIBILIDADE',
            'DESCRICAO_FORNECEDOR','ANALISE_COMPATIBILIDADE','COMPATIBILITY_SCORE','LAST_UPDATE'
        ]
        df_final = df_existing.reindex(columns=output_columns)

        output_dir = os.path.dirname(CAMINHO_SAIDA)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        try:
            writer = pd.ExcelWriter(CAMINHO_SAIDA, engine='openpyxl')
            df_final.to_excel(writer, index=False, sheet_name='Proposta')

            workbook = writer.book
            worksheet = writer.sheets['Proposta']

            for row_idx in range(2, len(df_final) + 2):
                score = df_final.at[row_idx - 2, 'COMPATIBILITY_SCORE']
                color_fill = get_rainbow_color(score)
                for col_idx in range(1, len(output_columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = color_fill

            writer.close()
            logger.info(f"Incremental save after processing item {item_index_in_df + 1}/{total_new_items}")
            print(f"Incremental save completed for item {item_index_in_df + 1}/{total_new_items}.")
        except Exception as e:
            logger.error(f"Failed to save Excel file: {e}")
            print(f"❌ - Failed to save Excel file: {e}")

    logger.info("All new items processed and saved incrementally.")
    print("✅ All new items processed and saved incrementally.")

if __name__ == "__main__":
    main()