import pandas as pd
import google.generativeai as genai
import google.api_core.exceptions as google_exceptions
import os
from dotenv import load_dotenv
import json
import time
from openpyxl.styles import PatternFill

# ======================================================================
# CONFIGURAÇÕES E CONSTANTES
# ======================================================================

# --- File Paths ---
BASE_DIR = r"C:\Users\pietr\OneDrive\.vscode\arte_"
CAMINHO_EDITAL = r"C:\Users\pietr\OneDrive\.vscode\arte_\DOWNLOADS\master.xlsx"
CAMINHO_SAIDA = r"C:\Users\pietr\OneDrive\.vscode\arte_\DOWNLOADS\master_heavy.xlsx"
CAMINHO_BASE = r"DOWNLOADS/RESULTADO_metadados/categoria_claude_sonnet_v5.xlsx"

# --- Financial Parameters ---
PROFIT_MARGIN = 0.53  # MARGEM DE LUCRO 
INITIAL_PRICE_FILTER_PERCENTAGE = 0.60  # FILTRO DE PREÇO DOS PRODUTOS NA BASE

# --- AI Model Configuration ---
# Lista de modelos para fallback. O script tentará o primeiro e, em caso de
# erro de cota, passará para o próximo da lista.
LLM_MODELS_FALLBACK = [
    "gemini-2.0-flash-lite",
    "gemini-2.5-flash",
    "gemini-2.5-flash-lite",
    "gemini-2.0-flash",
    "gemini-2.5-pro",
    "gemini-1.5-flash", 
]

# --- Subcategorization Keywords ------------------------------------
# A IA agora usará esta lista para identificar a subcategoria diretamente.
ALL_SUBCATEGORIES = [
    'arco', 'cavalete', 'corda', 'kit nut', 'kit rastilho', 'bag', 'banco', 
    'carrinho prancha', 'estante de partitura', 'suporte', 'baqueta', 
    'carrilhão', 'esteira', 'Máquina de Hi Hat', 'Pad para Bumbo', 'parafuso', 
    'pedal de bumbo', 'pele', 'prato', 'sino', 'talabarte', 'triângulo', 
    'graxa', 'oleo lubrificante', 'palheta de saxofone/clarinete', 
    'fone de ouvido', 'globo microfone', 'Interface de guitarra', 'pedal', 
    'mesa de som', 'microfone', 'cabo CFTV', 'cabo de rede', 'caixa medusa', 
    'Medusa', 'P10', 'P2xP10', 'painel de conexão', 'xlr M/F', 'amplificador', 
    'caixa de som', 'cubo para guitarra', "violino", "viola", "violão", 
    "guitarra", "baixo", "violoncelo", "afuché", "bateria", "bombo", "bumbo", 
    "caixa de guerra", "caixa tenor", "ganza", "pandeiro", "quadriton", 
    "reco reco", "surdo", "tambor", "tarol", "timbales", "trompete", 
    "bombardino", "trompa", "trombone", "tuba", "sousafone", "clarinete", 
    "saxofone", "flauta", "tuba bombardão", "flugelhorn", "euphonium", 
    "piano", "teclado digital", "glockenspiel", "metalofone"
]

# ============================================================ 
# FUNÇÕES DE SUPORTE
# ============================================================ 

def gerar_conteudo_com_fallback(prompt: str, modelos: list[str]) -> str | None:
    """
    Tenta gerar conteúdo usando uma lista de modelos em ordem de preferência.
    Se um modelo falhar por cota (ResourceExhausted), tenta o próximo.
    """
    for nome_modelo in modelos:
        try:
            print(f"   - Tentando chamada à API com o modelo: {nome_modelo}...")
            model = genai.GenerativeModel(nome_modelo)
            response = model.generate_content(prompt)

            if not response.parts:
                finish_reason = response.candidates[0].finish_reason.name if response.candidates else 'N/A'
                print(f"   - ❌ A GERAÇÃO RETORNOU VAZIA. Motivo: {finish_reason}. Isso pode ser causado por filtros de segurança.")
                return None

            print(f"   - Sucesso com o modelo '{nome_modelo}'.")
            return response.text
        except google_exceptions.ResourceExhausted as e:
            print(f"- ⚠️ Cota excedida para o modelo '{nome_modelo}'. Tentando o próximo da lista.")
            time.sleep(5)
            continue
        except Exception as e:
            print(f"   - ❌ Erro inesperado com o modelo '{nome_modelo}': {e}")
            return None
    print("   - ❌ FALHA TOTAL: Todos os modelos na lista de fallback falharam.")
    return None

def get_subcategory_from_ai(description: str, subcategories: list) -> str:
    """Usa o modelo de IA para extrair a subcategoria mais apropriada da descrição do item."""
    print("-🗣️ Asking AI for the item subcategory...")
    
    prompt = f"""
    <objetivo>
    Você é um especialista em instrumentos musicais e equipamentos de áudio.
    Sua tarefa é analisar a descrição do item a seguir e identificar a qual subcategoria de produto ele pertence.
    Responda APENAS com o nome da subcategoria da lista fornecida que melhor descreve o item.
    Se o item parecer uma combinação ou não se encaixar perfeitamente, escolha a subcategoria do componente principal. 
    </objetivo>

    <item_descricao>
    {description}
    </item_descricao>

    <lista_de_subcategorias_validas>
    {json.dumps(subcategories, indent=2)}
    </lista_de_subcategorias_validas>

    Subcategoria:
    """
    
    response_text = gerar_conteudo_com_fallback(prompt, LLM_MODELS_FALLBACK)
    if response_text:
        # Limpa a resposta para garantir que seja apenas a subcategoria
        subcategory = response_text.strip().replace("'", "").replace('"', '').lower()
        
        # Validação para garantir que a IA retornou uma subcategoria válida da lista
        if subcategory in [s.lower() for s in subcategories]:
            # Retorna a subcategoria com a capitalização original da lista
            original_case_subcategory = next(s for s in subcategories if s.lower() == subcategory)
            return original_case_subcategory
        else:
            print(f"   - WARNING: AI returned an invalid or new subcategory: '{subcategory}'. Defaulting to OUTROS.")
            return "OUTROS"
    else:
        print(f"   - ERROR: AI call for subcategorization failed for all models.")
        return "OUTROS"


def get_best_match_from_ai(item_edital, df_candidates):
    """Usa o modelo de IA para encontrar o melhor match dentro dos candidatos."""
    print("-🗣️Asking AI for the best match...")

    candidates_json = df_candidates[['DESCRICAO','categoria_principal','subcategoria','MARCA','MODELO','VALOR']] \
                        .to_json(orient="records", force_ascii=False, indent=2)

    prompt = f"""<identidade>Você é um consultor de licitações com 20+ anos de experiência em áudio/instrumentos, focado na Lei 14.133/21, economicidade e menor preço.</identidade>
<objetivo>
1. Analise tecnicamente o item do edital: "{item_edital['DESCRICAO']}"
2. Compare-o com cada produto na <base_fornecedores_filtrada>.
3. **Seleção Primária**: Encontre o produto da base que seja >=95% compatível. Dentre os compatíveis, escolha o de **menor 'Valor'**.
4. **Seleção Secundária**: Se nenhum produto for >=95% compatível, identifique o produto tecnicamente mais próximo e detalhe as especificações que não foram atendidas.
5. Use o Google para pesquisar e confirmar as especificações técnicas dos produtos candidatos para garantir uma análise precisa.
6. Responda **apenas** com um objeto JSON seguindo o <formato_saida>.
</objetivo>

<formato_saida>
Responda APENAS com um único objeto JSON. Não inclua ```json, explicações ou qualquer outro texto.

**CASO 1: Encontrou um produto >=95% compatível.**
{{
  "best_match": {{
    "Marca": "Marca do Produto",
    "Modelo": "Modelo do Produto",
    "Valor": 1234.56,
    "Descricao_fornecedor": "Descrição completa do produto na base",
    "Compatibilidade_analise": "Análise quantitativa da compatibilidade. Ex: 'Compatível. Atende a 5 de 6 especificações chave.'"
  }}
}}

**CASO 2: NENHUM produto é >=95% compatível.**
Retorne "best_match" como `null`. Adicionalmente, inclua "closest_match" com os dados do produto mais próximo e "reasoning" explicando o motivo da incompatibilidade.
{{
  "best_match": null,
  "reasoning": "Explique o principal motivo da incompatibilidade. Ex: 'Nenhum produto atende à especificação de material X ou potência Y.'",
  "closest_match": {{
    "Marca": "Marca do Produto Mais Próximo",
    "Modelo": "Modelo do Mais Próximo",
    "Valor": 4321.98,
    "Descricao_fornecedor": "Descrição do produto mais próximo.",
    "Compatibilidade_analise": "Análise da compatibilidade parcial com detalhes. Ex: 'Atende [especificação A, B], mas falha em [especificação C (material), D (potência)].'"
  }}
}}
</formato_saida>

<base_fornecedores_filtrada>
{candidates_json}
</base_fornecedores_filtrada>"""

    response_text = gerar_conteudo_com_fallback(prompt, LLM_MODELS_FALLBACK)
    if response_text:
        try:
            cleaned_response = response_text.strip().replace("```json","",).replace("```","")
            return json.loads(cleaned_response)
        except json.JSONDecodeError as e:
            print(f"   - ERROR decoding JSON from AI: {e}")
            return {"best_match": None, "closest_match": None, "reasoning": f"Erro na decodificação do JSON da API: {e}"}
    else:
        return {"best_match": None, "closest_match": None, "reasoning": "Falha na chamada da API para todos os modelos de fallback."}

# ============================================================ 
# MAIN
# ============================================================ 

def main():
    print("Starting the product matching process...")

    load_dotenv()
    api_key = os.getenv("GOOGLE_API_KEY")
    if not api_key:
        print("ERROR: GOOGLE_API_KEY not found in .env file.")
        return

    genai.configure(api_key=api_key)

    try:
        df_edital = pd.read_excel(CAMINHO_EDITAL)
        df_base = pd.read_excel(CAMINHO_BASE)
        print(f"👾 Edital loaded: {len(df_edital)} items from {os.path.basename(CAMINHO_EDITAL)}")
        print(f"📯 Product base loaded: {len(df_base)} products from {os.path.basename(CAMINHO_BASE)}")
    except FileNotFoundError as e:
        print(f"ERROR: Could not load data files. Details: {e}")
        return

    if os.path.exists(CAMINHO_SAIDA):
        print(f"   - Loading existing data from: {os.path.basename(CAMINHO_SAIDA)}")
        df_existing = pd.read_excel(CAMINHO_SAIDA)
        existing_keys = set(zip(df_existing['ARQUIVO'].astype(str), df_existing['Nº'].astype(str)))
    else:
        df_existing = pd.DataFrame()
        existing_keys = set()

    df_edital_new = df_edital[~df_edital.apply(lambda row: (str(row['ARQUIVO']), str(row['Nº'])) in existing_keys, axis=1)].copy()

    if df_edital_new.empty:
        print("\n✅ No new items to process. The output file is already up to date.")
        return

    print(f"   - Found {len(df_edital_new)} new items to process.")
    results = []
    total_new_items = len(df_edital_new)

    for idx, item_edital in df_edital_new.iterrows():
        descricao = str(item_edital['DESCRICAO'])
        print(f"\n 📈 Processing new item {df_edital_new.index.get_loc(idx) + 1}/{total_new_items}: {descricao[:60]}...")
        status = ""
        best_match_data = None
        closest_match_data = None
        reasoning = None
        data_to_populate = None

        valor_unit_edital = float(str(item_edital.get('VALOR_UNIT', '0')).replace(',', '.'))

        if valor_unit_edital > 0:
            max_cost = valor_unit_edital * INITIAL_PRICE_FILTER_PERCENTAGE
            df_candidates = df_base[df_base['VALOR'] <= max_cost].copy()
        else:
            print("- ℹ️ Valor de referência do edital é R$0.00 ou inválido. Analisando todos os produtos da base.")
            df_candidates = df_base.copy()

        if df_candidates.empty:
            if valor_unit_edital > 0:
                print(f"- ⚠️ Nenhum produto encontrado abaixo do custo máximo de R${max_cost:.2f}.")
            status = "Nenhum Produto com Margem"
        else:
            # --- LÓGICA DE FILTRAGEM ATUALIZADA ---
            # 1. Obter a subcategoria específica via IA
            item_subcategory = get_subcategory_from_ai(descricao, ALL_SUBCATEGORIES)
            time.sleep(10)
            print(f"   - AI categorized item into subcategory: {item_subcategory}")

            # 2. Filtrar a base de dados pela subcategoria retornada
            df_final_candidates = df_candidates[df_candidates['subcategoria'].str.lower() == item_subcategory.lower()]

            if df_final_candidates.empty:
                print(f"- ⚠️ Nenhum produto encontrado na subcategoria '{item_subcategory}'.")
                status = "Nenhum Produto na Subcategoria"
            else:
                print(f"  🦆 - Temos {len(df_final_candidates)} candidatos após a filtragem por subcategoria.")
                ai_result = get_best_match_from_ai(item_edital, df_final_candidates)
                time.sleep(30)

                best_match_data = ai_result.get("best_match")
                closest_match_data = ai_result.get("closest_match")
                reasoning = ai_result.get("reasoning")

                if best_match_data:
                    print(f" ✅ - AI recomenda: {best_match_data.get('Marca', 'N/A')} {best_match_data.get('Modelo', 'N/A')}")
                    status = "Match Encontrado"
                    data_to_populate = best_match_data
                elif closest_match_data:
                    print(f"- 💍 AI sugere como mais próximo: {closest_match_data.get('Marca', 'N/A')} {closest_match_data.get('Modelo', 'N/A')}")
                    if reasoning:
                        print(f"    Motivo: {reasoning}")
                    status = "Match Parcial (Sugestão)"
                    data_to_populate = closest_match_data
                else:
                    print(" ❌ - AI não encontrou produto compatível ou próximo.")
                    if reasoning:
                        print(f"    Motivo: {reasoning}")
                    status = "Nenhum Produto Compatível"


        result_row = {
            'ARQUIVO': item_edital['ARQUIVO'],
            'Nº': item_edital['Nº'],
            'DESCRICAO_EDITAL': item_edital['DESCRICAO'],
            'VALOR_UNIT_EDITAL': item_edital['VALOR_UNIT'],
            'STATUS': status,
            'QTDE': item_edital.get('QTDE', 0),
            'LOCAL_ENTREGA': item_edital.get('LOCAL_ENTREGA', ''),
            'INTERVALO_LANCES': item_edital.get('INTERVALO_LANCES', ''),
            'MOTIVO_INCOMPATIBILIDADE': reasoning if status != "Match Encontrado" else None
        }

        if data_to_populate:
            cost_price = float(data_to_populate.get('Valor') or 0)
            final_price = cost_price * (1 + PROFIT_MARGIN)
            margem_valor = final_price - cost_price
            qtde = int(item_edital.get('QTDE', 0) or 0)
            margem_total = margem_valor * qtde
            result_row.update({
                'MARCA_SUGERIDA': data_to_populate.get('Marca'),
                'MODELO_SUGERIDO': data_to_populate.get('Modelo'),
                'CUSTO_FORNECEDOR': cost_price,
                'PRECO_FINAL_VENDA': final_price,
                'MARGEM_LUCRO_VALOR': margem_valor,
                'MARGEM_LUCRO_TOTAL': margem_total,
                'DESCRICAO_FORNECEDOR': data_to_populate.get('Descricao_fornecedor'),
                'ANALISE_COMPATIBILIDADE': data_to_populate.get('Compatibilidade_analise')
            })

        results.append(result_row)

    print("\nProcess finished. Generating output file...")

    if not results:
        print("- ⚠️ No new results were generated to add.")
        return

    df_results = pd.DataFrame(results)
    
    df_final = pd.concat([df_existing, df_results], ignore_index=True)

    output_columns = [
        'ARQUIVO','Nº','STATUS','DESCRICAO_EDITAL','VALOR_UNIT_EDITAL',
        'MARCA_SUGERIDA','MODELO_SUGERIDO','QTDE','CUSTO_FORNECEDOR',
        'PRECO_FINAL_VENDA','MARGEM_LUCRO_VALOR','MARGEM_LUCRO_TOTAL',
        'MOTIVO_INCOMPATIBILIDADE','DESCRICAO_FORNECEDOR','ANALISE_COMPATIBILIDADE',
        'LOCAL_ENTREGA','INTERVALO_LANCES'
    ]
    
    for col in output_columns:
        if col not in df_final.columns:
            df_final[col] = ''
    df_final = df_final[output_columns]

    output_dir = os.path.dirname(CAMINHO_SAIDA)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    try:
        writer = pd.ExcelWriter(CAMINHO_SAIDA, engine='openpyxl')
        df_final.to_excel(writer, index=False, sheet_name='Proposta')

        workbook = writer.book
        worksheet = writer.sheets['Proposta']

        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        for row_idx, status in enumerate(df_final['STATUS'], 2):
            fill_to_apply = None
            if status == "Match Encontrado":
                fill_to_apply = green_fill
            elif status == "Match Parcial (Sugestão)":
                fill_to_apply = yellow_fill
            elif status in ["Nenhum Produto com Margem", "Nenhum Produto na Subcategoria", "Nenhum Produto Compatível"]:
                fill_to_apply = red_fill

            if fill_to_apply:
                for col_idx in range(1, len(df_final.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = fill_to_apply
        
        writer.close()
        print(f"✅ Success! Output file updated at: {CAMINHO_SAIDA}")
    except Exception as e:
        print(f"❌ Error writing output file: {e}")


if __name__ == "__main__":
    main()
