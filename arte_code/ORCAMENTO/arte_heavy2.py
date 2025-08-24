import pandas as pd
import google.generativeai as genai
import os
from dotenv import load_dotenv
import json
import time
from openpyxl.styles import PatternFill

# ======================================================================
# CONFIGURAÇÕES E CONSTANTES
# ======================================================================

# --- File Paths ---
CAMINHO_EDITAL = r"sheets/EDITAL/master.xlsx"
CAMINHO_BASE = r'C:\Users\pietr\OneDrive\.vscode\arte_\sheets\RESULTADO_metadados\produtos_categorizados_v3.xlsx'
CAMINHO_SAIDA = "sheets/RESULTADO_proposta/master_proposta_corzinha.xlsx"

# --- Financial Parameters ---
PROFIT_MARGIN = 0.53  # MARGEM DE LUCRO 
INITIAL_PRICE_FILTER_PERCENTAGE = 0.60  # FILTRO DE PREÇO DOS PRODUTOS NA BASE

# --- AI Model Configuration ---
GEMINI_MODEL = "gemini-2.5-flash-lite"

# --- Categorization Keywords ------------------------------------
CATEGORIZATION_KEYWORDS = {
    'ACESSORIO_CORDA': ['arco','cavalete','corda','kit nut','kit rastilho'],
    'ACESSORIO_GERAL': ['bag','banco','carrinho prancha','estante de partitura','suporte'],
    'ACESSORIO_PERCURSSAO': ['baqueta','carrilhão','esteira','Máquina de Hi Hat','Pad para Bumbo','parafuso','pedal de bumbo','pele','prato','sino','talabarte','triângulo'],
    'ACESSORIO_SOPRO': ['graxa','oleo lubrificante','palheta de saxofone/clarinete'],
    'EQUIPAMENTO_AUDIO': ['fone de ouvido','globo microfone','Interface de guitarra','pedal','mesa de som','microfone'],
    'EQUIPAMENTO_CABO': ['cabo CFTV','cabo de rede','caixa medusa','Medusa','P10','P2xP10','painel de conexão','xlr M/F'],
    'EQUIPAMENTO_SOM': ['amplificador','caixa de som','cubo para guitarra'],
    "INSTRUMENTO_CORDA": ["violino","viola","violão","guitarra","baixo","violoncelo"],
    "INSTRUMENTO_PERCUSSAO": ["afuché","bateria","bombo","bumbo","caixa de guerra","caixa tenor","ganza","pandeiro","quadriton","reco reco","surdo","tambor","tarol","timbales"],
    "INSTRUMENTO_SOPRO": ["trompete","bombardino","trompa","trombone","tuba","sousafone","clarinete","saxofone","flauta","tuba bombardão","flugelhorn","euphonium"],
    "INSTRUMENTO_TECLAS": ["piano","teclado digital","glockenspiel","metalofone"],
}

def load_categorized_products(file_path):
    """Carrega produtos categorizados de uma planilha Excel."""
    try:
        df = pd.read_excel(file_path)
        return df.set_index('ID_PRODUTO')[['categoria_principal', 'subcategoria']].to_dict('index')
    except FileNotFoundError:
        print(f"ERROR: Could not load categorized products file: {file_path}")
        return {}

# ============================================================
# FUNÇÕES DE SUPORTE
# ============================================================

def categorize_item(model, description: str, categories: list) -> str:
    """Usa o modelo de IA para classificar o item em uma das categorias fornecidas."""
    print(" - 🪼 - Asking AI for the item category...")
    
    prompt = f"""
    <objetivo>
    Você é um especialista em instrumentos musicais e equipamentos de áudio.
    Sua tarefa é classificar o item a seguir na categoria mais apropriada de uma lista fornecida.
    Responda APENAS com o nome da categoria escolhida.
    </objetivo>

    <item_descricao>
    {description}
    </item_descricao>

    <lista_de_categorias>
    {json.dumps(categories, indent=2)}
    </lista_de_categorias>

    Categoria:
    """
    
    try:
        response = model.generate_content(prompt)
        category = response.text.strip().replace("'", "").replace('"', '')
        if category in categories:
            return category
        else:
            print(f"   - WARNING: AI returned an invalid category: '{category}'. Defaulting to OUTROS.")
            return "OUTROS"
    except Exception as e:
        print(f"   - ERROR during category AI call: {e}")
        return "OUTROS"



def get_best_match_from_ai(model, item_edital, df_candidates):
    """Usa o modelo de IA para encontrar o melhor match dentro dos candidatos."""
    print(" - 🪼 - Asking AI for the best match...")

    candidates_json = df_candidates[['DESCRICAO','categoria_principal','subcategoria','MARCA','MODELO','VALOR']] \
                        .to_json(orient="records", force_ascii=False, indent=2)

    prompt = f"""<identidade>Você é um consultor de licitações com 20+ anos de experiência em áudio/instrumentos, focado na Lei 14.133/21, economicidade e menor preço.</identidade>
<objetivo>
1. Analise tecnicamente o item do edital: "{item_edital['DESCRICAO']}"
2. Compare-o com cada produto na <base_fornecedores_filtrada>.
3. **Seleção Primária**: Encontre o produto da base que seja >=95% compatível. Dentre os compatíveis, escolha o de **menor 'Valor'**.
4. **Seleção Secundária**: Se nenhum produto for >=95% compatível, identifique o produto tecnicamente mais próximo e detalhe as especificações que não foram atendidas.
5. Use o Google para pesquisar e confirmar as especificações técnicas dos produtos candidatos para garantir uma análise precisa.
6. Responda **apenas** com um objeto JSON seguindo o <formato_saida>.
</objetivo>

<formato_saida>
Responda APENAS com um único objeto JSON. Não inclua ```json, explicações ou qualquer outro texto.

**CASO 1: Encontrou um produto >=95% compatível.**
{{
  "best_match": {{
    "Marca": "Marca do Produto",
    "Modelo": "Modelo do Produto",
    "Valor": 1234.56,
    "Descricao_fornecedor": "Descrição completa do produto na base",
    "Compatibilidade_analise": "Análise quantitativa da compatibilidade. Ex: 'Compatível. Atende a 5 de 6 especificações chave.'"
  }}
}}

**CASO 2: NENHUM produto é >=95% compatível.**
Retorne "best_match" como `null`. Adicionalmente, inclua "closest_match" com os dados do produto mais próximo e "reasoning" explicando o motivo da incompatibilidade.
{{
  "best_match": null,
  "reasoning": "Explique o principal motivo da incompatibilidade. Ex: 'Nenhum produto atende à especificação de material X ou potência Y.'",
  "closest_match": {{
    "Marca": "Marca do Produto Mais Próximo",
    "Modelo": "Modelo do Mais Próximo",
    "Valor": 4321.98,
    "Descricao_fornecedor": "Descrição do produto mais próximo.",
    "Compatibilidade_analise": "Análise da compatibilidade parcial com detalhes. Ex: 'Atende [especificação A, B], mas falha em [especificação C (material), D (potência)].'"
  }}
}}
</formato_saida>

<base_fornecedores_filtrada>
{candidates_json}
</base_fornecedores_filtrada>"""

    try:
        response = model.generate_content(prompt)
        cleaned_response = response.text.strip().replace("```json","").replace("```","")
        return json.loads(cleaned_response)
    except Exception as e:
        print(f"   - ERROR during AI call: {e}")
        return {"best_match": None, "closest_match": None, "reasoning": f"Erro na API ou na análise da resposta: {e}"}

# ============================================================
# MAIN
# ============================================================

def main():
    print("Starting the product matching process...")

    load_dotenv()
    api_key = os.getenv("GOOGLE_API_KEY")
    if not api_key:
        print("ERROR: GOOGLE_API_KEY not found in .env file.")
        return

    genai.configure(api_key=api_key)
    model = genai.GenerativeModel(GEMINI_MODEL)

    try:
        df_edital = pd.read_excel(CAMINHO_EDITAL)
        df_base = pd.read_excel(CAMINHO_BASE)
        print(f"👾 Edital loaded: {len(df_edital)} items")
        print(f"📯 Product base loaded: {len(df_base)} products")
    except FileNotFoundError as e:
        print(f"ERROR: Could not load data files. Details: {e}")
        return

    results = []
    total_items = len(df_edital)

    for idx, item_edital in df_edital.iterrows():
        descricao = str(item_edital['DESCRICAO'])
        print(f"\n 📈 Processando item {idx + 1}/{total_items}: {descricao[:60]}...")
        status = ""
        best_match_data = None
        closest_match_data = None
        reasoning = None
        data_to_populate = None

        valor_unit_edital = item_edital['VALOR_UNIT']
        max_cost = valor_unit_edital * INITIAL_PRICE_FILTER_PERCENTAGE
        df_candidates = df_base[df_base['VALOR'] <= max_cost].copy()

        if df_candidates.empty:
            print(f"- ⚠️ - No products found below the max cost of R${max_cost:.2f}.")
            status = "Nenhum Produto com Margem"
        else:
            item_category = categorize_item(model, descricao, list(CATEGORIZATION_KEYWORDS.keys()))
            time.sleep(10) # To avoid rate limiting
            print(f"   - AI categorized item as: {item_category}")
            df_final_candidates = df_candidates[df_candidates['categoria_principal'] == item_category]

            if df_final_candidates.empty:
                print(f"- ⚠️ - Nenhum produto encontrado na categoria '{item_category}'.")
                status = "Nenhum Produto na Categoria"
            else:
                print(f"  🦆 - Temos {len(df_final_candidates)} candidatos depois da filtragem.")
                ai_result = get_best_match_from_ai(model, item_edital, df_final_candidates)
                time.sleep(30)

                best_match_data = ai_result.get("best_match")
                closest_match_data = ai_result.get("closest_match")
                reasoning = ai_result.get("reasoning")

                if best_match_data:
                    print(f" ✅ - AI recomenda: {best_match_data.get('Marca', 'N/A')} {best_match_data.get('Modelo', 'N/A')}")
                    status = "Match Encontrado"
                    data_to_populate = best_match_data
                elif closest_match_data:
                    print(f" ⚠️ - AI sugere como mais próximo: {closest_match_data.get('Marca', 'N/A')} {closest_match_data.get('Modelo', 'N/A')}")
                    if reasoning:
                        print(f"    Motivo: {reasoning}")
                    status = "Match Parcial (Sugestão)"
                    data_to_populate = closest_match_data
                else:
                    print(" ❌ - AI não encontrou produto compatível ou próximo.")
                    if reasoning:
                        print(f"    Motivo: {reasoning}")
                    status = "Nenhum Produto Compatível"

        result_row = {
            'ARQUIVO': item_edital['ARQUIVO'],
            'Nº': item_edital['Nº'],
            'DESCRICAO_EDITAL': item_edital['DESCRICAO'],
            'VALOR_UNIT_EDITAL': item_edital['VALOR_UNIT'],
            'STATUS': status,
            'MOTIVO_INCOMPATIBILIDADE': reasoning if status != "Match Encontrado" else None
        }

        if data_to_populate:
            cost_price = float(data_to_populate.get('Valor') or 0)
            final_price = cost_price * (1 + PROFIT_MARGIN)
            result_row.update({
                'MARCA_SUGERIDA': data_to_populate.get('Marca'),
                'MODELO_SUGERIDO': data_to_populate.get('Modelo'),
                'CUSTO_FORNECEDOR': cost_price,
                'PRECO_FINAL_VENDA': final_price,
                'MARGEM_LUCRO_VALOR': final_price - cost_price,
                'DESCRICAO_FORNECEDOR': data_to_populate.get('Descricao_fornecedor'),
                'ANALISE_COMPATIBILIDADE': data_to_populate.get('Compatibilidade_analise')
            })

        results.append(result_row)

    print("\nProcess finished. Generating output file...")

    if results:
        df_results = pd.DataFrame(results)
        output_columns = [
            'ARQUIVO','Nº','STATUS','DESCRICAO_EDITAL','VALOR_UNIT_EDITAL',
            'MARCA_SUGERIDA','MODELO_SUGERIDO','CUSTO_FORNECEDOR',
            'PRECO_FINAL_VENDA','MARGEM_LUCRO_VALOR','MOTIVO_INCOMPATIBILIDADE',
            'DESCRICAO_FORNECEDOR','ANALISE_COMPATIBILIDADE'
        ]
        for col in output_columns:
            if col not in df_results.columns:
                df_results[col] = ''
        df_results = df_results[output_columns]

        output_dir = os.path.dirname(CAMINHO_SAIDA)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # --- Geração do Excel com cores ---
        writer = pd.ExcelWriter(CAMINHO_SAIDA, engine='openpyxl')
        df_results.to_excel(writer, index=False, sheet_name='Proposta')

        workbook = writer.book
        worksheet = writer.sheets['Proposta']

        # Definição das cores de preenchimento
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Verde claro
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid') # Amarelo
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')      # Vermelho claro

        # Itera sobre as linhas do DataFrame para aplicar a formatação
        for row_idx, status in enumerate(df_results['STATUS'], 2):  # Começa em 2 para pular o cabeçalho
            fill_to_apply = None
            if status == "Match Encontrado":
                fill_to_apply = green_fill
            elif status == "Match Parcial (Sugestão)":
                fill_to_apply = yellow_fill
            elif status in ["Nenhum Produto com Margem", "Nenhum Produto na Categoria", "Nenhum Produto Compatível"]:
                fill_to_apply = red_fill

            if fill_to_apply:
                for col_idx in range(1, len(df_results.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = fill_to_apply
        
        writer.close()
        print(f"✅ Success! Output file generated at: {CAMINHO_SAIDA}")
    else:
        print("⚠️ No results were generated.")

if __name__ == "__main__":
    main()