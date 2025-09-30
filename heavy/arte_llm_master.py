import pandas as pd
import numpy as np
import os
import time
import json
import logging
from dotenv import load_dotenv
import joblib
import google.generativeai as genai
import google.api_core.exceptions as google_exceptions
import sys
import torch
from concurrent.futures import ThreadPoolExecutor, as_completed  # Para paralelização das chamadas LLM
from openpyxl.styles import PatternFill
from datetime import datetime
import torch

# Scikit-learn para ML clássico
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.ensemble import RandomForestClassifier
from sklearn.pipeline import Pipeline
from sklearn.metrics import accuracy_score, classification_report

# Sentence-Transformers para busca semântica
from sentence_transformers import SentenceTransformer, util

# =====================================================================
# CONFIGURAÇÕES E CONSTANTES
# =====================================================================

# --- File Paths ---
BASE_DIR = r"C:\Users\pietr\OneDrive\.vscode\arte_"  # Ajuste conforme seu ambiente
CAMINHO_EDITAL = os.path.join(BASE_DIR, "DOWNLOADS", "master.xlsx")
CAMINHO_BASE_PRODUTOS = os.path.join(BASE_DIR, "DOWNLOADS", "METADADOS", "produtos_metadados.xlsx")
CAMINHO_SAIDA = os.path.join(BASE_DIR, "DOWNLOADS", "master_otimizado.xlsx")

# --- ML Models & Precomputed Data Paths ---
MODELS_DIR = os.path.join(BASE_DIR, "machine_learning", "ml_models")
CLASSIFIER_PATH = os.path.join(MODELS_DIR, "subcategory_classifier.joblib")
EMBEDDINGS_PATH = os.path.join(MODELS_DIR, "product_embeddings.npy")
EMBEDDINGS_DATA_PATH = os.path.join(MODELS_DIR, "product_embeddings_data.pkl")
# Nova variável para o modelo Sentence Transformer
SENTENCE_TRANSFORMER_MODEL = 'paraphrase-multilingual-MiniLM-L12-v2'

# --- Financial Parameters ---
PROFIT_MARGIN = 0.47  # Margem de lucro
PRICE_FILTER_PERCENTAGE = 0.67  # Custo do fornecedor não pode exceder 75% do valor de referência do edital

# --- ML & AI Parameters ---
SEMANTIC_SEARCH_TOP_K = 15  # Nº de candidatos que a busca semântica vai levantar para o LLM (aumentado ligeiramente)
MIN_COMPATIBILITY_SCORE_FOR_MATCH = 90  # Score mínimo para ser considerado um "Match Encontrado"
CRITICAL_FAILURE_SCORE = 20.0  # Score máximo se houver uma "FALHA CRÍTICA"

# --- AI Model Configuration ---
# Usaremos Gemini pela robustez e estabilidade, como discutido.
LLM_MODELS_FALLBACK = [
    
    "gemini-flash-latest",
    "gemini-2.5-flash-lite",
    "gemini-2.0-flash-lite",
    "gemini-2.0-flash",
    "gemini-1.5-flash",
    "gemini-2.5-pro",
    "gemini-2.5-flash",
    
]
LLM_MAX_RETRIES = 3  # Tentativas para cada chamada LLM
LLM_TIMEOUT = 180  # Segundos
MAX_LLM_CONCURRENT_CALLS = 5  # Número de chamadas LLM que podem rodar em paralelo

# --- Logging Configuration ---
LOG_FILE = os.path.join(BASE_DIR, "LOGS", "arte_otimizado.log")
os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)  # Garante que a pasta de logs exista
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# =====================================================================
# FUNÇÕES DE IA (ADAPTADAS E OTIMIZADAS)
# =====================================================================

def gerar_conteudo_com_fallback(prompt: str, modelos: list[str]) -> str | None:
    """Tenta gerar conteúdo usando uma lista de modelos em ordem de preferência."""
    for nome_modelo in modelos:
        for attempt in range(LLM_MAX_RETRIES):
            try:
                logger.info(f"   - Tentando (tentativa {attempt+1}/{LLM_MAX_RETRIES}) modelo '{nome_modelo}'...")
                model = genai.GenerativeModel(nome_modelo)
                response = model.generate_content(prompt, request_options={"timeout": LLM_TIMEOUT})
                
                # Verifica se a resposta contém partes válidas
                if not response.parts:
                    finish_reason = response.candidates[0].finish_reason.name if response.candidates else 'N/A'
                    logger.warning(f"   - ❌ Resposta vazia do modelo '{nome_modelo}'. Motivo: {finish_reason}.")
                    if attempt == LLM_MAX_RETRIES - 1:  # Se for a última tentativa, falha
                        break  # Sai do loop de tentativas para este modelo
                    else:  # Se não for a última, tenta novamente
                        time.sleep(2 * (attempt + 1))  # Espera antes de re-tentar
                        continue
                logger.info(f"   - Sucesso com o modelo '{nome_modelo}'.")
                return response.text
            except google_exceptions.ResourceExhausted as e:
                logger.warning(f"- Cota excedida para o modelo '{nome_modelo}'. Tentando o próximo da lista.")
                time.sleep(5)
                break  # Sai do loop de tentativas e passa para o próximo modelo
            except Exception as e:
                logger.error(f"   - ❌ Erro inesperado com o modelo '{nome_modelo}': {e}")
                if attempt == LLM_MAX_RETRIES - 1:
                    break
                else:
                    time.sleep(2 * (attempt + 1))
                    continue
    logger.error("   - ❌ FALHA TOTAL: Todos os modelos e tentativas falharam.")
    return None

def get_best_match_from_ai(item_edital_dict: dict, df_candidates: pd.DataFrame):
    """
    Usa o modelo de IA para encontrar o melhor match e calcular um score de compatibilidade,
    seguindo o prompt detalhado discutido.
    """
    logger.info(f" - Enviando {len(df_candidates)} candidatos para validação final do LLM...")
    if df_candidates.empty:
        return {"best_match": None, "closest_match": None, "reasoning": "Nenhum candidato fornecido para o LLM."}

    # A IA espera a coluna 'DESCRICAO' para os candidatos
    df_candidates_ia = df_candidates.rename(columns={'DESCRICAO_FORNECEDOR': 'DESCRICAO'}).copy()  # .copy() para evitar SettingWithCopyWarning
    
    # Seleciona apenas colunas relevantes para o prompt
    candidates_json = df_candidates_ia[['DESCRICAO','subcategoria','MARCA','MODELO','VALOR']].to_json(orient="records", force_ascii=False, indent=2)
    
    # O item_edital_dict já deve ter as colunas 'DESCRICAO' e 'REFERENCIA'
    item_edital_json = json.dumps(item_edital_dict, ensure_ascii=False, indent=2)

    prompt = f"""<identidade>Você é um consultor de licitações com 20+ anos de experiência, especialista em análise técnica de produtos de áudio e instrumentos musicais. Seu foco é identificar a **compatibilidade técnica perfeita** (ou a mais próxima possível) com os requisitos do edital, garantindo a **economicidade** e o **menor preço** entre os produtos compatíveis, sempre aderindo aos princípios da Lei 14.133/21. Você é preciso, objetivo e detalhista na sua análise.</identidade>

<objetivo>
1.  **Análise Detalhada do Item do Edital:** Decomponha a descrição e referência do item do edital em **todas as suas especificações técnicas-chave e requisitos obrigatórios**. Identifique aspectos como tipo de produto, potência, dimensões, material, funcionalidades, acessórios inclusos, padrões de conectividade, etc.

2.  **Comparação Rigorosa com Candidatos:** Compare CADA produto fornecido na `<lista_produtos_candidatos>` com as especificações detalhadas do item do edital.

3.  **Prioridade: Compatibilidade Perfeita (>= 98%):**
    *   Busque um produto que atenda a **TODAS ou quase todas (>= 98%) as especificações técnicas e requisitos obrigatórios** do edital.
    *   Se encontrar múltiplos produtos com essa alta compatibilidade, selecione aquele com o **MENOR 'Valor'**. Este será o seu `best_match`.
    *   Para o `best_match`, gere um `Compatibilidade_analise` detalhada.

4.  **Fallback: Melhor Compatível (se não houver match perfeito):**
    *   Se nenhum produto atingir `>= 98%` de compatibilidade, identifique o produto que seja **tecnicamente o mais próximo e viável** (ainda que não perfeito). Este será o seu `closest_match`.
    *   Para o `closest_match`, gere um `Compatibilidade_analise` detalhada.

5.  **Identificação de Falhas Críticas:**
    *   Qualquer especificação ou requisito obrigatório do edital que **NÃO FOR ATENDIDO** pelo produto (seja `best_match` ou `closest_match`) deve ser claramente identificado nos `pontos_negativos` com o prefixo **`FALHA CRÍTICA:`**.
    *   Exemplos de falhas críticas: Potência mínima não atingida, material diferente do solicitado, ausência de funcionalidade essencial (ex: sensor térmico), marca ou modelo *específico* exigido não compatível.
    *   (Opcional: Se dúvida em specs/preços, valide via web search - integre tools como 'web_search' aqui se disponível).

6.  **Formato de Saída (APENAS JSON):** Retorne estritamente um único objeto JSON, seguindo os formatos abaixo. Não inclua ` ```json ` ou qualquer outro texto ao redor.

</objetivo>

<formato_saida>
// CASO 1: Encontrou um produto com compatibilidade >= 98%
{{
  "best_match": {{
    "Marca": "Marca do Produto",
    "Modelo": "Modelo do Produto",
    "Valor": 1234.56,
    "Descricao_fornecedor": "Descrição completa do produto na base",
    "Compatibilidade_analise": {{
      "score_proposto": 99, // Score de 0 a 100
      "pontos_positivos": [
        "Atende à potência de 50W RMS em 8 ohms.",
        "Material da carcaça em alumínio anodizado, conforme edital.",
        "Inclui software de controle e drivers atualizados."
      ],
      "pontos_negativos": [
        "Não inclui cabo P10 de 5 metros (acessório comum, pode ser adquirido à parte)."
      ],
      "justificativa": "Compatibilidade excelente, atende a todas as especificações técnicas críticas. A pequena ausência de um acessório comum não compromete a funcionalidade principal."
    }}
  }},
  "closest_match": null // DEVE ser null neste caso
}}

// CASO 2: NENHUM produto atingiu >= 98% de compatibilidade
{{
  "best_match": null, // DEVE ser null neste caso
  "closest_match": {{
    "Marca": "Marca do Produto Mais Próximo",
    "Modelo": "Modelo do Mais Próximo",
    "Valor": 4321.98,
    "Descricao_fornecedor": "Descrição do produto mais próximo na base.",
    "Compatibilidade_analise": {{
      "score_proposto": 70, // Score de 0 a 100
      "pontos_positivos": [
        "É da mesma categoria e potência aproximada (40W vs 50W solicitados).",
        "Oferece conectividade USB-C."
      ],
      "pontos_negativos": [
        "FALHA CRÍTICA: Não possui sensor térmico integrado, requisito obrigatório de segurança.",
        "FALHA CRÍTICA: Zoom óptico de 2x, enquanto o edital pede 56x.",
        "Material da carcaça em plástico, não em metal resistente como especificado."
      ],
      "justificativa": "Incompatível devido à ausência de funcionalidades críticas (sensor térmico e zoom adequado) e material de construção inferior. Não atende aos requisitos mínimos de segurança e desempenho."
    }}
  }}
}}
</formato_saida>

<item_edital>
{item_edital_json}
</item_edital>

<lista_produtos_candidatos>
{candidates_json}
</lista_produtos_candidatos>

JSON:
"""
    response_text = gerar_conteudo_com_fallback(prompt, LLM_MODELS_FALLBACK)
    if response_text:
        try:
            cleaned_response = response_text.strip().replace("```json", "").replace("```", "").strip()
            return json.loads(cleaned_response)
        except json.JSONDecodeError as e:
            logger.error(f"   - ERRO decodificando JSON da IA: {e}. Resposta bruta: {response_text[:500]}...")
            return {"best_match": None, "closest_match": None, "reasoning": f"Erro na decodificação do JSON da API: {e}"}
    return {"best_match": None, "closest_match": None, "reasoning": "Falha na chamada da API para todos os modelos."}

# =====================================================================
# FUNÇÕES DE MACHINE LEARNING (OTIMIZADAS)
# =====================================================================

def train_classifier(df_path: str, classifier_save_path: str):
    """
    Treina um modelo de classificação (Random Forest) para prever a 'subcategoria'
    a partir da 'DESCRICAO' e salva o pipeline treinado.
    """
    logger.info("Iniciando treinamento do classificador de subcategoria...")
    df = pd.read_excel(df_path)
    df.dropna(subset=['DESCRICAO', 'subcategoria'], inplace=True)
    df['DESCRICAO'] = df['DESCRICAO'].astype(str)
    df['subcategoria'] = df['subcategoria'].astype(str)

    X = df['DESCRICAO']
    y = df['subcategoria']

    if len(y.unique()) < 2:
        logger.error("Não há categorias suficientes (mínimo 2) para treinar o classificador. Abortando.")
        return None

    # Regra para divisão: estratificada se possível
    min_samples_per_class = y.value_counts().min()
    if min_samples_per_class < 2:
        logger.warning(f"Alguma subcategoria tem menos de 2 amostras ({min_samples_per_class}). Usando divisão sem estratificação.")
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
    else:
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42, stratify=y)

    # Pipeline otimizado
    classifier_pipeline = Pipeline([
        ('tfidf', TfidfVectorizer(sublinear_tf=True, min_df=5, ngram_range=(1, 2), stop_words='portuguese')),
        ('clf', RandomForestClassifier(n_estimators=100, random_state=42, class_weight='balanced'))
    ])

    logger.info("Treinando o pipeline...")
    classifier_pipeline.fit(X_train, y_train)

    # Avaliação
    y_pred = classifier_pipeline.predict(X_test)
    accuracy = accuracy_score(y_test, y_pred)
    logger.info(f"Acurácia do classificador no set de teste: {accuracy:.2f}")
    logger.info("Relatório de Classificação:\n" + classification_report(y_test, y_pred, zero_division=0))

    logger.info(f"Salvando o pipeline treinado em: {classifier_save_path}")
    joblib.dump(classifier_pipeline, classifier_save_path)
    logger.info("Treinamento concluído e modelo salvo.")
    return classifier_pipeline

def generate_embeddings(df_path: str, embeddings_save_path: str, data_save_path: str):
    """
    Gera e salva os embeddings de texto para todas as descrições de produtos
    usando um modelo SentenceTransformer. Inclui 'categoria_principal' se presente.
    """
    logger.info("Iniciando geração de embeddings semânticos para os produtos...")
    df = pd.read_excel(df_path)
    df.dropna(subset=['DESCRICAO'], inplace=True)
    df.rename(columns={'DESCRICAO': 'DESCRICAO_FORNECEDOR'}, inplace=True)  # Padroniza nome da coluna
    df['DESCRICAO_FORNECEDOR'] = df['DESCRICAO_FORNECEDOR'].astype(str)

    logger.info(f"Carregando o modelo Sentence Transformer '{SENTENCE_TRANSFORMER_MODEL}'...")
    model = SentenceTransformer(SENTENCE_TRANSFORMER_MODEL)

    logger.info(f"Gerando embeddings para {len(df)} descrições de produtos...")
    df['VALOR'] = pd.to_numeric(df['VALOR'], errors='coerce').fillna(0)
    
    embeddings = model.encode(df['DESCRICAO_FORNECEDOR'].tolist(), convert_to_tensor=True, show_progress_bar=True)

    logger.info(f"Salvando embeddings em: {embeddings_save_path}")
    np.save(embeddings_save_path, embeddings.cpu().numpy())
    
    # Salva apenas as colunas necessárias para o contexto de busca
    cols_to_save = ['DESCRICAO_FORNECEDOR', 'subcategoria', 'MARCA', 'MODELO', 'VALOR']
    if 'categoria_principal' in df.columns:
        cols_to_save.append('categoria_principal')
    
    df_to_save = df[[col for col in cols_to_save if col in df.columns]].copy()
    df_to_save.to_pickle(data_save_path)
    
    logger.info(f"Dados dos embeddings salvos em: {data_save_path}")
    logger.info("Geração de embeddings concluída.")

# =====================================================================
# FUNÇÕES AUXILIARES E DE PÓS-PROCESSAMENTO
# =====================================================================

def calculate_compatibility_score(analise_obj: dict) -> float:
    """
    Calcula a pontuação de compatibilidade (0-100) de forma estruturada,
    baseando-se no objeto JSON retornado pela IA, com prioridade para "FALHA CRÍTICA".
    """
    if not isinstance(analise_obj, dict):
        return 0.0

    score_proposto = analise_obj.get('score_proposto')
    pontos_negativos = analise_obj.get('pontos_negativos', [])

    # Regra 1: Penalização máxima para FALHAS CRÍTICAS
    if any('FALHA CRÍTICA:' in str(p).upper() for p in pontos_negativos):
        logger.debug(f"Falha crítica detectada, score fixado em {CRITICAL_FAILURE_SCORE}.")
        return CRITICAL_FAILURE_SCORE  # Score muito baixo para indicar inviabilidade

    # Regra 2: Usar o score proposto pela IA se for válido e sem falhas críticas
    if isinstance(score_proposto, (int, float)) and 0 <= score_proposto <= 100:
        logger.debug(f"Usando score proposto pela IA: {score_proposto}.")
        # Se a IA deu um score alto, mas ainda há pontos negativos, ajusta ligeiramente
        if score_proposto > MIN_COMPATIBILITY_SCORE_FOR_MATCH and pontos_negativos:
            return min(score_proposto, 95.0)  # Limita a 95 se não for perfeito, mas a IA deu alto
        return float(score_proposto)

    # Regra 3: Fallback quantitativo se o score proposto for inválido ou ausente
    pontos_positivos = analise_obj.get('pontos_positivos', [])
    total_pontos = len(pontos_positivos) + len(pontos_negativos)

    if total_pontos == 0:
        logger.debug("Nenhuma análise detalhada (pontos +/-). Score 0.")
        return 0.0  # Nenhuma análise detalhada

    total = len(pontos_positivos) + len(pontos_negativos)
    score = (len(pontos_positivos) / total) * 100
    logger.debug(f"Calculando score fallback: {score}.")
    
    # Ajuste final se o cálculo resultou em 100 mas há negativos (mesmo não críticos)
    if score == 100 and pontos_negativos:
        return 95.0

    return score

def get_rainbow_color(score: float) -> PatternFill:
    """Gera cor com base no score de compatibilidade."""
    if pd.isna(score):
        score = 0.0
    # Cores mais distintivas
    strong_green, light_green, yellow, orange, red = ('00B050', 'C6EFCE', 'FFEB9C', 'FFC000', 'FFC7CE')
    if score == 100: hex_color = strong_green
    elif score >= MIN_COMPATIBILITY_SCORE_FOR_MATCH: hex_color = light_green
    elif score >= 60: hex_color = yellow
    elif score >= CRITICAL_FAILURE_SCORE: hex_color = orange  # Laranja para scores médios/baixos mas não críticos
    else: hex_color = red  # Vermelho para scores muito baixos ou críticos
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def save_styled_excel(file_path: str, df: pd.DataFrame, output_cols: list):
    """
    Salva o DataFrame final em um arquivo Excel, aplicando formatação de cores
    na coluna 'COMPATIBILITY_SCORE'.
    """
    logger.info("Aplicando formatação de cores e salvando arquivo final...")
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Proposta')
            worksheet = writer.sheets['Proposta']
            
            # Adiciona filtro e congela painel
            worksheet.auto_filter.ref = worksheet.dimensions
            worksheet.freeze_panes = 'A2'

            for r_idx, row in df.iterrows():
                score = row.get('COMPATIBILITY_SCORE')
                fill = get_rainbow_color(score)
                # +2 porque Excel é 1-based e a primeira linha é o cabeçalho
                for c_idx in range(1, len(output_cols) + 1):
                    worksheet.cell(row=r_idx + 2, column=c_idx).fill = fill
        logger.info(f"✅ Arquivo final estilizado e salvo em: {file_path}")

    except Exception as e:
        logger.error(f"Ocorreu um erro ao estilizar e salvar o arquivo Excel: {e}", exc_info=True)

# =====================================================================
# PIPELINE PRINCIPAL DE PROCESSAMENTO DO ITEM
# =====================================================================

def process_item(item_edital_row, df_products_base, classifier, st_model, product_embeddings_data, product_embeddings_tensor, main_categories_list):
    """Processa um único item do edital através do pipeline de ML e LLM."""
    item_edital_dict = item_edital_row.to_dict()
    item_desc = str(item_edital_dict['DESCRICAO'])
    valor_unit_edital = pd.to_numeric(item_edital_dict.get('VALOR_UNIT'), errors='coerce')
    
    final_row_data = {
        **item_edital_dict,
        'DESCRICAO_EDITAL': item_edital_dict.get('DESCRICAO'),
        'VALOR_UNIT_EDITAL': valor_unit_edital,
        'STATUS': 'Nenhum Match Encontrado',
        'COMPATIBILITY_SCORE': 0.0,
        'LAST_UPDATE': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

    df_price_filtered = df_products_base.copy()
    if pd.notna(valor_unit_edital) and valor_unit_edital > 0:
        max_cost = valor_unit_edital * PRICE_FILTER_PERCENTAGE
        df_price_filtered = df_products_base[df_products_base['VALOR'] <= max_cost].copy()
        if df_price_filtered.empty:
            logger.warning(f"Item {item_edital_dict['Nº']}: Nenhum produto na base atende ao critério de preço inicial (custo <= R${max_cost:.2f}).")
            final_row_data['STATUS'] = 'Nenhum Produto com Margem'
            return final_row_data

    ai_result = {}
    best_match = None
    closest_match = None
    
    # --- ETAPA 1: Classificação ML da Subcategoria ---
    logger.info(f"   [ETAPA 1/4] Classificando subcategoria via ML...")
    predicted_subcategory = classifier.predict([item_desc])[0]
    logger.info(f"   - Subcategoria prevista pelo ML: '{predicted_subcategory}'")
    
    # --- ETAPA 2: Busca Semântica na Subcategoria Prevista + LLM ---
    logger.info("   [ETAPA 2/4] Buscando e validando LLM na subcategoria...")
    df_candidates_subcat = df_price_filtered[df_price_filtered['subcategoria'] == predicted_subcategory].copy()

    if not df_candidates_subcat.empty:
        item_embedding = st_model.encode(item_desc, convert_to_tensor=True)
        # Filtra os embeddings para apenas os candidatos relevantes
        candidate_indices_in_full_df = df_candidates_subcat.index.tolist()
        candidate_embeddings_for_search = product_embeddings_tensor[candidate_indices_in_full_df]  # Usa tensor diretamente

        cos_scores = util.cos_sim(item_embedding, candidate_embeddings_for_search)[0]
        top_k = min(SEMANTIC_SEARCH_TOP_K, len(df_candidates_subcat))
        
        # Obter os índices dos top_k no sub-DataFrame
        top_results_relative_indices = np.argpartition(-cos_scores.cpu().numpy(), range(top_k))[:top_k]
        
        # Mapear de volta para os índices originais do df_products_base
        final_candidate_indices = [candidate_indices_in_full_df[i] for i in top_results_relative_indices]
        df_llm_candidates = df_products_base.iloc[final_candidate_indices].copy()
        
        ai_result = get_best_match_from_ai(item_edital_dict, df_llm_candidates)
        best_match = ai_result.get("best_match")
        closest_match = ai_result.get("closest_match")
        
        if best_match:
            logger.info(f"   ✅ [ETAPA 2 SUCESSO] Match perfeito (>90%) encontrado na subcategoria ML: {best_match.get('Marca')} {best_match.get('Modelo')}")
            # Atualiza final_row_data com os dados do best_match
            cost_price = pd.to_numeric(best_match.get('Valor'), errors='coerce')
            final_price = cost_price * (1 + PROFIT_MARGIN) if pd.notna(cost_price) else 0.0
            qtde = pd.to_numeric(item_edital_dict.get('QTDE'), errors='coerce') or 0
            
            analise_compat_obj = best_match.get('Compatibilidade_analise', {})
            compat_score = calculate_compatibility_score(analise_compat_obj)
            
            result_row_update = {
                'STATUS': 'Match Encontrado',
                'MARCA_SUGERIDA': best_match.get('Marca'),
                'MODELO_SUGERIDO': best_match.get('Modelo'),
                'DESCRICAO_FORNECEDOR': best_match.get('Descricao_fornecedor'),
                'CUSTO_FORNECEDOR': cost_price,
                'PRECO_FINAL_VENDA': final_price,
                'COMPATIBILITY_SCORE': compat_score,
                'ANALISE_COMPATIBILIDADE': json.dumps(analise_compat_obj, ensure_ascii=False)
            }
            final_row_data.update(result_row_update)
            return final_row_data
    else:
        logger.warning(f"   - Nenhum produto em '{predicted_subcategory}' após filtro de preço.")

    # --- ETAPA 3: FALLBACK - Classificação da Categoria Principal (via LLM) + Busca Semântica + LLM ---
    # Só tenta se a busca na subcategoria não encontrou um match "perfeito"
    logger.info("   [ETAPA 3/4] FALLBACK: Classificando Categoria Principal via LLM e buscando...")
    item_desc_full = f"{item_edital_dict['DESCRICAO']} {item_edital_dict.get('REFERENCIA', '')}"
    
    # Este prompt é simplificado para ser mais rápido
    prompt_main_cat = f"""Analise a descrição do item e classifique-o em uma das seguintes categorias principais. Responda APENAS com o nome da categoria.

Item: \"{item_desc_full}\"

Categorias Principais Válidas: {main_categories_list}

Categoria Principal:"""
    correct_main_cat = gerar_conteudo_com_fallback(prompt_main_cat, LLM_MODELS_FALLBACK)
    if correct_main_cat and correct_main_cat.strip() in main_categories_list:
        correct_main_cat = correct_main_cat.strip()
        logger.info(f"   - Categoria principal prevista pelo LLM: '{correct_main_cat}'")
        df_candidates_maincat = df_price_filtered[df_price_filtered['categoria_principal'] == correct_main_cat].copy()

        if not df_candidates_maincat.empty:
            item_embedding = st_model.encode(item_desc, convert_to_tensor=True)
            candidate_indices_in_full_df_main = df_candidates_maincat.index.tolist()
            candidate_embeddings_for_search_main = product_embeddings_tensor[candidate_indices_in_full_df_main]

            cos_scores_main = util.cos_sim(item_embedding, candidate_embeddings_for_search_main)[0]
            top_k_main = min(SEMANTIC_SEARCH_TOP_K, len(df_candidates_maincat))
            top_results_relative_indices_main = np.argpartition(-cos_scores_main.cpu().numpy(), range(top_k_main))[:top_k_main]
            final_candidate_indices_main = [candidate_indices_in_full_df_main[i] for i in top_results_relative_indices_main]
            df_llm_candidates_main = df_products_base.iloc[final_candidate_indices_main].copy()
            
            ai_result_main_cat = get_best_match_from_ai(item_edital_dict, df_llm_candidates_main)
            best_match_main_cat = ai_result_main_cat.get("best_match")
            closest_match_main_cat = ai_result_main_cat.get("closest_match")

            if best_match_main_cat:
                logger.info(f"   ✅ [ETAPA 3 SUCESSO] Match perfeito (>90%) encontrado na categoria principal LLM: {best_match_main_cat.get('Marca')} {best_match_main_cat.get('Modelo')}")
                # Atualiza final_row_data com os dados do best_match_main_cat
                cost_price = pd.to_numeric(best_match_main_cat.get('Valor'), errors='coerce')
                final_price = cost_price * (1 + PROFIT_MARGIN) if pd.notna(cost_price) else 0.0
                qtde = pd.to_numeric(item_edital_dict.get('QTDE'), errors='coerce') or 0
                
                analise_compat_obj = best_match_main_cat.get('Compatibilidade_analise', {})
                compat_score = calculate_compatibility_score(analise_compat_obj)
                
                result_row_update = {
                    'STATUS': 'Match Encontrado',
                    'MARCA_SUGERIDA': best_match_main_cat.get('Marca'),
                    'MODELO_SUGERIDO': best_match_main_cat.get('Modelo'),
                    'DESCRICAO_FORNECEDOR': best_match_main_cat.get('Descricao_fornecedor'),
                    'CUSTO_FORNECEDOR': cost_price,
                    'PRECO_FINAL_VENDA': final_price,
                    'COMPATIBILITY_SCORE': compat_score,
                    'ANALISE_COMPATIBILIDADE': json.dumps(analise_compat_obj, ensure_ascii=False)
                }
                final_row_data.update(result_row_update)
                return final_row_data
            elif closest_match_main_cat:
                logger.info(f"   ⚠️ [ETAPA 3 PARCIAL] Closest match encontrado na categoria principal LLM: {closest_match_main_cat.get('Marca')} {closest_match_main_cat.get('Modelo')}")
                cost_price = pd.to_numeric(closest_match_main_cat.get('Valor'), errors='coerce')
                final_price = cost_price * (1 + PROFIT_MARGIN) if pd.notna(cost_price) else 0.0
                qtde = pd.to_numeric(item_edital_dict.get('QTDE'), errors='coerce') or 0
                
                analise_compat_obj = closest_match_main_cat.get('Compatibilidade_analise', {})
                compat_score = calculate_compatibility_score(analise_compat_obj)
                
                result_row_update = {
                    'STATUS': 'Match Parcial (Sugestão)',
                    'MARCA_SUGERIDA': closest_match_main_cat.get('Marca'),
                    'MODELO_SUGERIDO': closest_match_main_cat.get('Modelo'),
                    'DESCRICAO_FORNECEDOR': closest_match_main_cat.get('Descricao_fornecedor'),
                    'CUSTO_FORNECEDOR': cost_price,
                    'PRECO_FINAL_VENDA': final_price,
                    'COMPATIBILITY_SCORE': compat_score,
                    'ANALISE_COMPATIBILIDADE': json.dumps(analise_compat_obj, ensure_ascii=False)
                }
                final_row_data.update(result_row_update)
                return final_row_data
    else:
        logger.warning("   - LLM não retornou uma categoria principal válida.")

    # --- ETAPA 4: FALLBACK FINAL - Busca na Base Inteira + LLM ---
    logger.info("   [ETAPA 4/4] FALLBACK FINAL: Buscando e validando LLM na base inteira (pós-preço)...")
    item_embedding = st_model.encode(item_desc, convert_to_tensor=True)
    cos_scores_full = util.cos_sim(item_embedding, product_embeddings_tensor)[0]
    top_k_full = min(SEMANTIC_SEARCH_TOP_K * 2, len(df_price_filtered))  # Aumenta top_k para fallback
    top_results_full = np.argpartition(-cos_scores_full.cpu().numpy(), range(top_k_full))[:top_k_full]
    df_llm_candidates_full = df_products_base.iloc[top_results_full].copy()
    
    ai_result_full = get_best_match_from_ai(item_edital_dict, df_llm_candidates_full)
    best_match_full = ai_result_full.get("best_match")
    closest_match_full = ai_result_full.get("closest_match")

    if best_match_full:
        logger.info(f"   ✅ [ETAPA 4 SUCESSO] Match perfeito (>90%) encontrado na base inteira: {best_match_full.get('Marca')} {best_match_full.get('Modelo')}")
        cost_price = pd.to_numeric(best_match_full.get('Valor'), errors='coerce')
        final_price = cost_price * (1 + PROFIT_MARGIN) if pd.notna(cost_price) else 0.0
        qtde = pd.to_numeric(item_edital_dict.get('QTDE'), errors='coerce') or 0
        
        analise_compat_obj = best_match_full.get('Compatibilidade_analise', {})
        compat_score = calculate_compatibility_score(analise_compat_obj)
        
        result_row_update = {
            'STATUS': 'Match Encontrado (Fallback Base)',
            'MARCA_SUGERIDA': best_match_full.get('Marca'),
            'MODELO_SUGERIDO': best_match_full.get('Modelo'),
            'DESCRICAO_FORNECEDOR': best_match_full.get('Descricao_fornecedor'),
            'CUSTO_FORNECEDOR': cost_price,
            'PRECO_FINAL_VENDA': final_price,
            'COMPATIBILITY_SCORE': compat_score,
            'ANALISE_COMPATIBILIDADE': json.dumps(analise_compat_obj, ensure_ascii=False)
        }
        final_row_data.update(result_row_update)
    elif closest_match_full:
        logger.info(f"   ⚠️ [ETAPA 4 PARCIAL] Closest match encontrado na base inteira: {closest_match_full.get('Marca')} {closest_match_full.get('Modelo')}")
        cost_price = pd.to_numeric(closest_match_full.get('Valor'), errors='coerce')
        final_price = cost_price * (1 + PROFIT_MARGIN) if pd.notna(cost_price) else 0.0
        qtde = pd.to_numeric(item_edital_dict.get('QTDE'), errors='coerce') or 0
        
        analise_compat_obj = closest_match_full.get('Compatibilidade_analise', {})
        compat_score = calculate_compatibility_score(analise_compat_obj)
        
        result_row_update = {
            'STATUS': 'Match Parcial (Fallback Base)',
            'MARCA_SUGERIDA': closest_match_full.get('Marca'),
            'MODELO_SUGERIDO': closest_match_full.get('Modelo'),
            'DESCRICAO_FORNECEDOR': closest_match_full.get('Descricao_fornecedor'),
            'CUSTO_FORNECEDOR': cost_price,
            'PRECO_FINAL_VENDA': final_price,
            'COMPATIBILITY_SCORE': compat_score,
            'ANALISE_COMPATIBILIDADE': json.dumps(analise_compat_obj, ensure_ascii=False)
        }
        final_row_data.update(result_row_update)
    else:
        final_row_data['MOTIVO_INCOMPATIBILIDADE'] = ai_result_full.get('reasoning', 'Nenhum match viável encontrado na base inteira.')

    return final_row_data

# =====================================================================
# MAIN: EXECUÇÃO PRINCIPAL DO SCRIPT
# =====================================================================

def main():
    logger.info("Iniciando o pipeline otimizado para processamento de itens do edital...")
    load_dotenv()
    genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))

    # Carrega os dados
    try:
        df_edital = pd.read_excel(CAMINHO_EDITAL)
        df_products_base = pd.read_excel(CAMINHO_BASE_PRODUTOS)
    except FileNotFoundError as e:
        logger.critical(f"Erro ao carregar arquivos: {e}")
        sys.exit(1)

    # Define categorias principais (extraídas dos dados anexos)
    main_categories_list = list(df_products_base['categoria_principal'].unique())  # Extrai dinamicamente

    # Treina o classificador se não existir
    if not os.path.exists(CLASSIFIER_PATH):
        classifier = train_classifier(CAMINHO_BASE_PRODUTOS, CLASSIFIER_PATH)
    else:
        logger.info(f"Carregando classificador pré-treinado de: {CLASSIFIER_PATH}")
        classifier = joblib.load(CLASSIFIER_PATH)

    # Gera embeddings se não existirem
    if not os.path.exists(EMBEDDINGS_PATH) or not os.path.exists(EMBEDDINGS_DATA_PATH):
        generate_embeddings(CAMINHO_BASE_PRODUTOS, EMBEDDINGS_PATH, EMBEDDINGS_DATA_PATH)

    logger.info(f"Carregando embeddings pré-computados de: {EMBEDDINGS_PATH}")
    product_embeddings_np = np.load(EMBEDDINGS_PATH)
    product_embeddings_tensor = torch.from_numpy(product_embeddings_np)  # Converte para tensor

    product_embeddings_data = pd.read_pickle(EMBEDDINGS_DATA_PATH)  # Dados contextuais

    st_model = SentenceTransformer(SENTENCE_TRANSFORMER_MODEL)

    # Verifica se o arquivo de saída existe; carrega para processamento incremental
    if os.path.exists(CAMINHO_SAIDA):
        df_existing = pd.read_excel(CAMINHO_SAIDA)
        existing_keys = set(zip(df_existing['ARQUIVO'].astype(str), pd.to_numeric(df_existing['Nº'], errors='coerce').fillna(0).astype(int).astype(str)))
    else:
        df_existing = pd.DataFrame()
        existing_keys = set()

    df_novos_itens = df_edital[~df_edital.apply(lambda row: (str(row['ARQUIVO']), str(int(pd.to_numeric(row['Nº'], errors='coerce') or 0))) in existing_keys, axis=1)].copy()

    if df_novos_itens.empty:
        logger.info("Nenhum novo item para processar. Arquivo de saída atualizado.")
        return

    logger.info(f"Identificados {len(df_novos_itens)} novos itens para processar.")

    output_columns = [
        'ARQUIVO', 'Nº', 'DESCRICAO_EDITAL', 'REFERENCIA', 'STATUS', 'UNID_FORN', 'QTDE', 
        'VALOR_UNIT_EDITAL', 'VALOR_TOTAL', 'LOCAL_ENTREGA',
        'MARCA_SUGERIDA', 'MODELO_SUGERIDO', 'DESCRICAO_FORNECEDOR', 
        'CUSTO_FORNECEDOR', 'PRECO_FINAL_VENDA',  
        'COMPATIBILITY_SCORE', 'ANALISE_COMPATIBILIDADE', 'MOTIVO_INCOMPATIBILIDADE', 'LAST_UPDATE'
    ]

    # Processamento paralelo com salvamento incremental a cada item
    with ThreadPoolExecutor(max_workers=MAX_LLM_CONCURRENT_CALLS) as executor:
        future_to_row = {executor.submit(process_item, row, df_products_base, classifier, st_model, product_embeddings_data, product_embeddings_tensor, main_categories_list): row for _, row in df_novos_itens.iterrows()}
        
        processed_count = 0
        total_items = len(df_novos_itens)
        
        for future in as_completed(future_to_row):
            row_data = future_to_row[future]
            try:
                new_result = future.result()
                
                # Concatena o novo resultado ao dataframe existente
                df_existing = pd.concat([df_existing, pd.DataFrame([new_result])], ignore_index=True)
                
                processed_count += 1
                item_id = row_data.get('Nº', 'N/A')
                logger.info(f"✅ Item Nº {item_id} processado com sucesso. Progresso: {processed_count}/{total_items}.")

            except Exception as e:
                item_id = row_data.get('Nº', 'N/A')
                file_name = row_data.get('ARQUIVO', 'N/A')
                logger.error(f"❌ Erro ao processar item Nº {item_id} do arquivo {file_name}: {e}", exc_info=True)
                # Salva um placeholder de erro para não reprocessar
                error_result = {
                    **row_data.to_dict(),
                    'STATUS': 'ERRO DE PROCESSAMENTO',
                    'MOTIVO_INCOMPATIBILIDADE': str(e),
                    'LAST_UPDATE': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                df_existing = pd.concat([df_existing, pd.DataFrame([error_result])], ignore_index=True)
            
            finally:
                # Salva o arquivo Excel completo após cada tentativa (sucesso ou falha)
                df_to_save = df_existing.reindex(columns=output_columns)
                save_styled_excel(CAMINHO_SAIDA, df_to_save, output_columns)
                logger.info(f"💾 Arquivo de saída salvo incrementalmente.")


    logger.info("✅ Processamento de todos os itens concluído.")

if __name__ == "__main__":
    main()