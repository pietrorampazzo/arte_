import pandas as pd
import requests
import os
from dotenv import load_dotenv
import json
import logging
from openpyxl.styles import PatternFill
from datetime import datetime
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np

# ======================================================================
# CONFIGURAÇÕES E CONSTANTES
# ======================================================================

# --- File Paths ---
BASE_DIR = r"C:\Users\pietr\OneDrive\.vscode\arte_"
CAMINHO_EDITAL = os.path.join(BASE_DIR, "DOWNLOADS", "master.xlsx")
CAMINHO_BASE = os.path.join(BASE_DIR, "DOWNLOADS", "METADADOS", "produtos_metadados.xlsx")
CAMINHO_SAIDA = os.path.join(BASE_DIR, "DOWNLOADS", "master_machine_open.xlsx")
CAMINHO_HEAVY_EXISTENTE = CAMINHO_SAIDA

# --- Financial Parameters ---
PROFIT_MARGIN = 0.53
INITIAL_PRICE_FILTER_PERCENTAGE = 0.60
EXPANDED_PRICE_FILTER_PERCENTAGE = 0.75

# --- AI Model Configuration (OpenRouter) ---
LLM_MODELS_FALLBACK = [

    "x-ai/grok-4-fast:free",
    'nvidia/nemotron-nano-9b-v2:free',
    'deepseek/deepseek-chat-v3.1:free',
    'openai/gpt-oss-120b:free',
    'openai/gpt-oss-20b:free',
    'z-ai/glm-4.5-air:free',
    'qwen/qwen3-coder:free',
    'moonshotai/kimi-k2:free',
    'google/gemma-3n-e2b-it:free',
    'tngtech/deepseek-r1t2-chimera:free',
    'mistralai/mistral-small-3.2-24b-instruct:free',
    'moonshotai/kimi-dev-72b:free',
    'deepseek/deepseek-r1-0528-qwen3-8b:free',
    'qwen/qwen3-8b:free',
    "mistralai/mistral-7b-instruct",
    "meta-llama/llama-3-8b-instruct",
]

# --- ML Configuration ---
SUBCATEGORY_ML_CANDIDATES = 10
MAIN_CATEGORY_ML_CANDIDATES = 20

# --- Categorization Keywords ---
CATEGORIZATION_KEYWORDS = {
        "EQUIPAMENTO_SOM" : ["caixa_ativa", "caixa_passiva", "caixa_portatil", "line_array", "subwoofer_ativo", "subwoofer_passivo", "amplificador_potencia", "cabeçote_amplificado", "coluna_vertical", "monitor_de_palco"],
        "EQUIPAMENTO_AUDIO" : ["microfone_dinamico", "microfone_condensador", "microfone_lapela", "microfone_sem_fio", "microfone_instrumento", "mesa_analogica", "mesa_digital", "interface_audio", "processador_dsp", "fone_monitor", "sistema_iem", "pedal_efeitos"],
        "INSTRUMENTO_CORDA" : ["violao", "guitarra", "contra_baixo", "violino", "violoncelo", "ukulele", "cavaquinho"],
        "INSTRUMENTO_PERCUSSAO" : ["bateria_acustica", "bateria_eletronica", "repinique", "rocari", "tantan", "rebolo","surdo_mao", "cuica", "zabumba", "caixa_guerra", "bombo_fanfarra", "lira_marcha","tarol", "malacacheta", "caixa_bateria", "pandeiro", "tamborim","reco_reco", "agogô", "triangulo", "chocalho", "afuche", "cajon", "bongo", "conga", "djembé", "timbal", "atabaque", "berimbau","tam_tam", "caxixi", "carilhao", "xequerê", "prato"],
        "INSTRUMENTO_SOPRO" : ["saxofone", "trompete", "trombone", "trompa", "clarinete", "flauta", "tuba", "flugelhorn", "bombardino", "corneta", "cornetão"],
        "INSTRUMENTO_TECLADO" : ["teclado_digital", "piano_acustico", "piano_digital", "sintetizador", "controlador_midi", "glockenspiel", "metalofone"],
        "ACESSORIO_MUSICAL" : ["banco_teclado", "estante_partitura", "suporte_microfone", "suporte_instrumento", "carrinho_transporte", "case_bag", "afinador", "metronomo", "cabos_audio", "palheta", "cordas", "oleo_lubrificante", "graxa", "surdina", "bocal_trompete", "pele_percussao", "baqueta", "talabarte", "pedal_bumbo", "chimbal_hihat"],
        "EQUIPAMENTO_TECNICO" : ["ssd", "fonte_energia", "switch_rede", "projetor", "drone"]
}

# --- Logging Configuration ---
LOG_FILE = os.path.join(BASE_DIR, "LOGS", "arte_stanley.log")
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ============================================================ 
# FUNÇÕES DE IA (OPENROUTER) E ML
# ============================================================ 

def gerar_conteudo_com_fallback(prompt: str, modelos: list[str]) -> str | None:
    """Tenta gerar conteúdo usando a API OpenRouter com uma lista de modelos de fallback."""
    api_key = os.getenv("OPENROUTER_API_KEY")
    if not api_key:
        logger.error("OPENROUTER_API_KEY não encontrada no arquivo .env")
        print("   - ❌ OPENROUTER_API_KEY não encontrada no arquivo .env")
        return None

    for nome_modelo in modelos:
        try:
            print(f"   - Tentando chamada à API OpenRouter com o modelo: {nome_modelo}...")
            
            response = requests.post(
                url="https://openrouter.ai/api/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "HTTP-Referer": "https://github.com/seu-usuario/arte-heavy",
                    "X-Title": "Arte Heavy Analysis",
                },
                data=json.dumps({
                    "model": nome_modelo,
                    "messages": [{"role": "user", "content": prompt}]
                }),
                timeout=180
            )

            if response.status_code == 429: # Rate limit
                print(f"   - Rate limit excedido para o modelo '{nome_modelo}'. Tentando o próximo.")
                continue
            
            response.raise_for_status()
            response_json = response.json()
            
            content = response_json.get('choices', [{}])[0].get('message', {}).get('content')
            if not content:
                logger.warning(f"API retornou resposta vazia para o modelo {nome_modelo}.")
                return None

            print(f"   - Sucesso com o modelo '{nome_modelo}'.")
            return content

        except requests.exceptions.RequestException as e:
            logger.error(f"Erro de requisição com o modelo '{nome_modelo}': {e}")
            continue
        except Exception as e:
            logger.critical(f"Erro inesperado e crítico com o modelo '{nome_modelo}': {e}")
            return None

    logger.error("FALHA TOTAL: Todos os modelos na lista de fallback falharam.")
    return None

def get_item_classification(description: str, reference: str, categories_with_subcategories: dict) -> dict | None:
    """Usa o modelo de IA para classificar o item."""
    print("- Asking AI for item classification...")
    prompt = f'''<identidade>Você é um especialista de almoxarifado e perito em catalogação de produtos.</identidade>
<objetivo>
Sua tarefa é classificar o item a seguir, identificando sua `categoria_principal` e `subcategoria` com base na estrutura fornecida.
- A `categoria_principal` DEVE ser uma das chaves da estrutura.
- A `subcategoria` DEVE ser um dos valores da lista associada à categoria principal escolhida.
Responda APENAS com um objeto JSON.
</objetivo>
<item_a_ser_classificado>
Descrição: {description}
Referência: {reference}
</item_a_ser_classificado>
<estrutura_de_categorias_e_subcategorias_permitidas>
{json.dumps(categories_with_subcategories, indent=2, ensure_ascii=False)}
</estrutura_de_categorias_e_subcategorias_permitidas>
<formato_saida>
{{
  "categoria_principal": "NOME_DA_CATEGORIA_PRINCIPAL",
  "subcategoria": "NOME_DA_SUBCATEGORIA_ESPECIFICA"
}}
</formato_saida>
JSON:
'''
    response_text = gerar_conteudo_com_fallback(prompt, LLM_MODELS_FALLBACK)
    if response_text:
        try:
            cleaned_response = response_text.strip().replace("```json", "").replace("```", "")
            classification = json.loads(cleaned_response)
            if 'categoria_principal' in classification and 'subcategoria' in classification and classification['categoria_principal'] in categories_with_subcategories:
                return classification
            else:
                logger.warning(f"AI returned an invalid classification: {classification}")
                return None
        except json.JSONDecodeError as e:
            logger.error(f"ERROR decoding JSON from AI for classification: {e}")
            return None
    return None

def get_best_match_from_ai(item_edital, df_candidates, attempt_description: str):
    """Usa o modelo de IA para encontrar o melhor match dentro dos candidatos."""
    print(f" - Asking AI for the best match ({attempt_description})...")
    if df_candidates.empty:
        return {"best_match": None, "closest_match": None, "reasoning": "No candidates provided."}

    candidates_json = df_candidates[['DESCRICAO','categoria_principal','subcategoria','MARCA','MODELO','VALOR']].to_json(orient="records", force_ascii=False, indent=2)
    prompt = f'''<identidade>Você é um consultor de licitações com 20+ anos de experiência em áudio/instrumentos, focado na Lei 14.133/21, economicidade e menor preço.</identidade>
<objetivo>
1.  Analise tecnicamente o item do edital: Descrição: `"{item_edital['DESCRICAO']}"` Referência: `"{item_edital.get('REFERENCIA', 'N/A')}"`.
2.  Compare-o com cada produto na `<base_fornecedores_filtrada>`.
3.  **Seleção Primária**: Encontre o produto da base que seja >=95% compatível. Dentre os compatíveis, escolha o de **menor 'Valor'**.
4.  **Seleção Secundária**: Se nenhum for >=95% compatível, identifique o produto tecnicamente mais próximo.
5.  **Análise Estruturada**: Para o produto escolhido (seja `best_match` ou `closest_match`), forneça uma análise de compatibilidade detalhada no formato JSON especificado abaixo.
6.  **Falhas Críticas**: Se uma especificação chave e obrigatória do edital não for atendida, adicione a flag `FALHA CRÍTICA:` no início do ponto negativo correspondente.
7.  Responda **apenas** com um objeto JSON.
</objetivo>
<formato_saida>
Responda APENAS com um único objeto JSON. Não inclua ```json ou qualquer outro texto.

**CASO 1: Encontrou um produto >=95% compatível.**
{{
  "best_match": {{
    "Marca": "Marca do Produto",
    "Modelo": "Modelo do Produto",
    "Valor": 1234.56,
    "Descricao_fornecedor": "Descrição completa do produto na base",
    "Compatibilidade_analise": {{
      "score_proposto": 98,
      "pontos_positivos": ["Atende à potência de 50W"],
      "pontos_negativos": ["Não inclui o cabo de força (acessório menor)"],
      "justificativa": "Excelente compatibilidade."
    }}
  }}
}}

**CASO 2: NENHUM produto é >=95% compatível.**
{{
  "best_match": null,
  "closest_match": {{
    "Marca": "Marca do Produto Mais Próximo",
    "Modelo": "Modelo do Mais Próximo",
    "Valor": 4321.98,
    "Descricao_fornecedor": "Descrição do produto mais próximo.",
    "Compatibilidade_analise": {{
      "score_proposto": 40,
      "pontos_positivos": ["É da mesma marca"],
      "pontos_negativos": ["FALHA CRÍTICA: Não possui sensor térmico"],
      "justificativa": "Incompatível devido à ausência de funcionalidades críticas."
    }}
  }}
}}
</formato_saida><base_fornecedores_filtrada>{candidates_json}</base_fornecedores_filtrada>'''

    response_text = gerar_conteudo_com_fallback(prompt, LLM_MODELS_FALLBACK)
    if response_text:
        try:
            cleaned_response = response_text.strip().replace("```json", "").replace("```", "")
            return json.loads(cleaned_response)
        except json.JSONDecodeError as e:
            logger.error(f"ERROR decoding JSON from AI (Attempt: {attempt_description}): {e}")
            return {"best_match": None, "closest_match": None, "reasoning": f"Erro na decodificação do JSON da API: {e}"}
    return {"best_match": None, "closest_match": None, "reasoning": "Falha na chamada da API para todos os modelos de fallback."}

def get_top_n_ml_matches(item_edital, df_candidates, n):
    if df_candidates.empty or n == 0:
        return pd.DataFrame()
    print(f" - Running ML to find top {n} matches...")
    edital_text = f"{item_edital['DESCRICAO']} {item_edital.get('REFERENCIA', '')}"
    candidates_texts = df_candidates['DESCRICAO'].fillna('').tolist()
    if not candidates_texts:
        return pd.DataFrame()
    vectorizer = TfidfVectorizer()
    all_texts = [edital_text] + candidates_texts
    tfidf_matrix = vectorizer.fit_transform(all_texts)
    similarities = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:]).flatten()
    num_results = min(n, len(similarities))
    top_indices = np.argsort(similarities)[-num_results:][::-1]
    print(f"   - Found {len(top_indices)} ML candidates.")
    return df_candidates.iloc[top_indices]

# ============================================================ 
# FUNÇÕES AUXILIARES E DE PROCESSAMENTO
# ============================================================ 

def check_match_found(ai_result: dict) -> bool:
    best_match_data = ai_result.get("best_match")
    if not best_match_data or not isinstance(best_match_data, dict):
        return False
    return isinstance(best_match_data.get('Compatibilidade_analise'), dict)

def calculate_compatibility_score(analise_obj: dict | str) -> float:
    if not isinstance(analise_obj, dict): return 0.0
    if any('FALHA CRÍTICA' in str(p).upper() for p in analise_obj.get('pontos_negativos', [])): return 20.0
    score_proposto = analise_obj.get('score_proposto')
    if isinstance(score_proposto, (int, float)) and 0 <= score_proposto <= 100: return float(score_proposto)
    return 0.0

def get_rainbow_color(score: float) -> PatternFill:
    if pd.isna(score): score = 0.0
    if score == 100: hex_color = '00B050'
    elif score >= 80: hex_color = 'C6EFCE'
    elif score >= 40: hex_color = 'FFEB9C'
    else: hex_color = 'FFC7CE'
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def process_single_item_pipeline(item_edital, df_base, price_filter_percentage, classification):
    """Executa o pipeline de análise para um único item do edital."""
    # CORREÇÃO: Inicializa ai_result para garantir que a variável sempre exista.
    ai_result = {"best_match": None, "closest_match": None, "reasoning": "Nenhum candidato viável encontrado no pipeline."}
    
    valor_unit_edital = float(str(item_edital.get('VALOR_UNIT', '0')).replace(',', '.'))
    df_price_candidates = df_base[df_base['VALOR'] <= (valor_unit_edital * price_filter_percentage)] if valor_unit_edital > 0 else df_base.copy()

    if df_price_candidates.empty: return "Nenhum Produto com Margem", None, None

    if not classification:
        print("   - ⚠️ AI classification failed. Cannot proceed with category filters.")
        df_ml_candidates = get_top_n_ml_matches(item_edital, df_price_candidates, MAIN_CATEGORY_ML_CANDIDATES)
        ai_result = get_best_match_from_ai(item_edital, df_ml_candidates, "Fallback ML on all price-filtered")
        return ("Match Encontrado (Fallback ML)" if check_match_found(ai_result) else "Nenhum Produto Compatível"), ai_result.get("best_match"), ai_result.get("closest_match")

    main_category, subcategory = classification.get('categoria_principal'), classification.get('subcategoria')
    print(f"   - AI classified item as: Categoria='{main_category}', Subcategoria='{subcategory}'")

    # Pipeline de tentativas
    for scope, df_filter, ml_candidates_n, attempt_desc in [
        ("Subcategoria", df_price_candidates[df_price_candidates['subcategoria'].str.contains(subcategory, case=False, na=False)], 0, "Subcategory Filter"),
        ("ML na Subcategoria", df_price_candidates[df_price_candidates['subcategoria'].str.contains(subcategory, case=False, na=False)], SUBCATEGORY_ML_CANDIDATES, f"ML Top {SUBCATEGORY_ML_CANDIDATES} in Subcategory"),
        ("ML na Categoria Principal", df_price_candidates[df_price_candidates['categoria_principal'] == main_category], MAIN_CATEGORY_ML_CANDIDATES, f"ML Top {MAIN_CATEGORY_ML_CANDIDATES} in Main Category")
    ]:
        if df_filter.empty: continue
        
        candidates = get_top_n_ml_matches(item_edital, df_filter, ml_candidates_n) if ml_candidates_n > 0 else df_filter
        if candidates.empty: continue

        ai_result = get_best_match_from_ai(item_edital, candidates, attempt_desc)
        if check_match_found(ai_result):
            return f"Match Encontrado ({scope})", ai_result.get("best_match"), None

    return "Nenhum Match >95%", ai_result.get("best_match"), ai_result.get("closest_match")

# ============================================================ 
# MAIN
# ============================================================ 

def main():
    logger.info("Starting the Stanley product matching process with OpenRouter (Incremental Save Mode)...")
    print("Starting the Stanley product matching process with OpenRouter (Incremental Save Mode)...")

    load_dotenv()
    if not os.getenv("OPENROUTER_API_KEY"):
        logger.critical("OPENROUTER_API_KEY not found in .env file.")
        print("ERRO: OPENROUTER_API_KEY não encontrada no arquivo .env. Por favor, adicione-a.")
        return

    try:
        df_edital = pd.read_excel(CAMINHO_EDITAL)
        df_base = pd.read_excel(CAMINHO_BASE)
        df_base['VALOR'] = pd.to_numeric(df_base['VALOR'], errors='coerce').fillna(0)
    except FileNotFoundError as e:
        logger.critical(f"Could not load data files: {e}")
        return

    try:
        df_existing = pd.read_excel(CAMINHO_HEAVY_EXISTENTE)
        existing_keys = set(zip(df_existing['ARQUIVO'].astype(str), pd.to_numeric(df_existing['Nº'], errors='coerce').fillna(0).astype(int).astype(str)))
    except FileNotFoundError:
        df_existing = pd.DataFrame()
        existing_keys = set()

    df_edital_new = df_edital[~df_edital.apply(lambda row: (str(row['ARQUIVO']), str(int(pd.to_numeric(row['Nº'], errors='coerce') or 0))) in existing_keys, axis=1)].copy()

    if df_edital_new.empty:
        logger.info("No new items to process. Output file is up to date.")
        print("\n✅ No new items to process. The output file is already up to date.")
        return

    logger.info(f"Identified {len(df_edital_new)} new items to process.")
    print(f"   - Found {len(df_edital_new)} new items to process.")
    
    total_new_items = len(df_edital_new)

    for i, item_edital in df_edital_new.iterrows():
        print(f"\n📈 Processing new item {i + 1}/{total_new_items}: {str(item_edital['DESCRICAO'])[:60]}...")

        classification = get_item_classification(
            str(item_edital['DESCRICAO']),
            str(item_edital.get('REFERENCIA', 'N/A')),
            CATEGORIZATION_KEYWORDS
        )

        status, best_match, closest_match = process_single_item_pipeline(
            item_edital, df_base, INITIAL_PRICE_FILTER_PERCENTAGE, classification
        )

        if "Match Encontrado" not in status:
            print("\n===== TENTATIVA 2: Filtro de Preço Expandido (75%) =====")
            logger.warning(f"Item {item_edital['Nº']} não encontrou match. Tentando com filtro de preço expandido.")
            status_exp, best_match_exp, closest_match_exp = process_single_item_pipeline(
                item_edital, df_base, EXPANDED_PRICE_FILTER_PERCENTAGE, classification
            )
            if "Match Encontrado" in status_exp or (closest_match_exp and not closest_match):
                status, best_match, closest_match = status_exp, best_match_exp, closest_match_exp

        data_to_populate = best_match if best_match else closest_match
        if not best_match and closest_match: status = "Match Parcial (Sugestão)"

        result_row = { **item_edital.to_dict(), 'STATUS': status, 'LAST_UPDATE': datetime.now().strftime('%Y-%m-%d %H:%M:%S') }

        if data_to_populate:
            cost_price = float(data_to_populate.get('Valor') or 0)
            final_price = cost_price * (1 + PROFIT_MARGIN)
            qtde = pd.to_numeric(item_edital.get('QTDE'), errors='coerce') or 0
            analise_compat_obj = data_to_populate.get('Compatibilidade_analise')
            compat_score = calculate_compatibility_score(analise_compat_obj)
            
            analise_text = ""
            if isinstance(analise_compat_obj, dict):
                justificativa = analise_compat_obj.get('justificativa', 'N/A')
                positivos = "; ".join(analise_compat_obj.get('pontos_positivos', []))
                negativos = "; ".join(analise_compat_obj.get('pontos_negativos', []))
                analise_text = f"Justificativa: {justificativa} | Prós: {positivos} | Contras: {negativos}"

            result_row.update({
                'MARCA_SUGERIDA': data_to_populate.get('Marca'), 'MODELO_SUGERIDO': data_to_populate.get('Modelo'),
                'CUSTO_FORNECEDOR': cost_price, 'PRECO_FINAL_VENDA': final_price,
                'LUCRO_TOTAL': (final_price - cost_price) * qtde,
                'DESCRICAO_FORNECEDOR': data_to_populate.get('Descricao_fornecedor'),
                'ANALISE_COMPATIBILIDADE': analise_text, 'COMPATIBILITY_SCORE': compat_score
            })
        
        # Lógica de salvamento incremental
        df_existing = pd.concat([df_existing, pd.DataFrame([result_row])], ignore_index=True)

        output_columns = [
            'ARQUIVO','Nº','DESCRICAO_EDITAL','REFERENCIA','STATUS',
            'UNID_FORN', 'QTDE', 'VALOR_UNIT_EDITAL', 'VALOR_TOTAL',
            'LOCAL_ENTREGA', 'INTERVALO_LANCES',
            'MARCA_SUGERIDA', 'MODELO_SUGERIDO', 'CUSTO_FORNECEDOR',
            'PRECO_FINAL_VENDA','MARGEM_LUCRO_VALOR', 'LUCRO_TOTAL', 'MOTIVO_INCOMPATIBILIDADE',
            'DESCRICAO_FORNECEDOR','ANALISE_COMPATIBILIDADE','COMPATIBILITY_SCORE','LAST_UPDATE'
        ]
        for col in output_columns:
            if col not in df_existing.columns:
                df_existing[col] = None
        df_final = df_existing[output_columns]

        try:
            with pd.ExcelWriter(CAMINHO_SAIDA, engine='openpyxl') as writer:
                df_final.to_excel(writer, index=False, sheet_name='Proposta')
                worksheet = writer.sheets['Proposta']
                for r_idx, row in df_final.iterrows():
                    fill = get_rainbow_color(row['COMPATIBILITY_SCORE'])
                    for c_idx in range(1, len(output_columns) + 1):
                        worksheet.cell(row=r_idx + 2, column=c_idx).fill = fill
            logger.info(f"Incremental save for item {i+1}/{total_new_items} successful.")
            print(f"  💾 Item {i+1} salvo com sucesso.")
        except Exception as e:
            logger.error(f"Failed to save Excel file incrementally: {e}", exc_info=True)
            print(f"  ❌ Erro ao salvar o item {i+1}: {e}")

    logger.info("All new items processed.")
    print("\n✅ All new items processed.")

if __name__ == "__main__":
    main()